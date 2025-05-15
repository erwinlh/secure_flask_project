import pandas as pd
import openpyxl
import mysql.connector
import requests
import xlwt
import io
from termcolor import colored
from utils.common import formatRut, convertir_fecha_iso_a_fecha_hora, convertir_datetime_a_fecha_iso, connect, insertar_resultados, ejecutar_consulta, decodificar_xml
from utils.services import actualizar_estado_sii_masivo, consultaEstado, consultaPDF, recuperar_xml, diccionario_a_payload_cliente, consultar_cliente, guardar_cliente, diccionario_a_payload_factura, guardar_venta, consultar_venta

rutEmisor = '76708884-1'
tipoDte = 33
folioDte = 310

# Define symbols
ERROR_SYMBOL = " ✗ "
OK_SYMBOL = " ✓ "
ASK_SYMBOL = " ? "
NEW_SYMBOL = " # "
WARNING_SYMBOL = " ! "

def exportar_consulta_xls(desde, hasta):
    """
    Función que ejecuta una consulta SQL y exporta el resultado a un archivo .xls.
    """
    try:# Generar un nombre de archivo único
        #timestamp = time.strftime("%Y%m%d_%H%M%S")
        #random_suffix = random.randint(100, 999)
        filename = f"GdeBoletas_{desde}_{hasta}.xls"
        #filename = f"GdeBoletas_{timestamp}_{desde}_{hasta}.xls"

        print(f"Iniciando exportación para {desde} a {hasta}. Archivo: {filename}")

        conexion = connect()
        #query = f"SELECT * FROM boletas WHERE fecha BETWEEN '{desde}' AND '{hasta}'"
        query = f"""select 'A' AS `Destino del Documento`,
        'N' AS `Documento impreso se genera nulo`,
        'N' AS `Documento impreso Genera Rebaja Stock`,
        (case when (`f`.`tipo` = 33) then 'FEAV' 
            when (`f`.`tipo` = 39) then 'BEAV' 
            when (`f`.`tipo` = 110) then 'FEAV' else NULL end) AS `Tipo de Documento`,
            `f`.`folio` AS `Número`
            ,'' AS `Número final, sólo boletas`,
            date_format(`f`.`fecha`,'%d-%m-%Y') AS `Fecha`,
            'Marina Dunas' AS `Local`,
            'MarinaDunas' AS `Vendedor`,
            'Peso' AS `Moneda Referencia`,
            1 AS `Tasa Referencia`,
            'CREDITO30' AS `Condición de Pago`,
            date_format((`f`.`fecha` + interval 30 day),'%d-%m-%Y') AS `Fecha de Vencimiento`,(case when (`f`.`tipo` = 33) then `f`.`rut` when (`f`.`tipo` = 110) then '55.555.555-5' when (`f`.`tipo` = 39) then '66.666.666-6' else NULL end) AS `Código de Cliente`,(case when (`f`.`tipo` = 33) then '21_FACTURASPORCOBRAR' when (`f`.`tipo` = 110) then '22_FACT_EXPPORCOBRAR' when (`f`.`tipo` = 39) then '23_BOLETASPORCOBRAR' else NULL end) AS `Tipo de Cliente`,'' AS `Centro de Negocios`,'' AS `Clasificador 1`,'' AS `Clasificador 2`,'' AS `Origen del Documento`,1 AS `Lista de Precio`,'' AS `Código del Proyecto`,`f`.`neto` AS `Afecto`,`f`.`exento` AS `Exento`,`f`.`total` AS `Total`,'' AS `Bodega Inventario`,'' AS `Motivo de movimiento Inventario`,'' AS `Centro de Negocios Inventario`,'' AS `Tipo de Cuenta Inventario`,'' AS `Proveedor Inventario`,'' AS `Dirección de Despacho`,'' AS `Clasificador 1 Inventario`,'' AS `Clasificador 2 Inventario`,(case when (`f`.`tipo` = 33) then `f`.`rut` when (`f`.`tipo` = 110) then '55.555.555-5' when (`f`.`tipo` = 39) then '66.666.666-6' else NULL end) AS `Código Legal`,(case when (`f`.`tipo` = 33) then `f`.`razon_social` when (`f`.`tipo` = 110) then 'CLIENTE EXTRANGERO' when (`f`.`tipo` = 39) then 'CLIENTE PARTICULAR' else NULL end) AS `Nombre`,'' AS `Giro`,'' AS `Dirección`,'' AS `Ciudad`,1 AS `Rubro`,concat('Conf: ',`f`.`conf_no`,' - ',`f`.`razon_social`) AS `Glosa`,1 AS `Línea de Detalle`,'S' AS `Articulo / Servicio`,'Alojamiento' AS `Código del Producto`,1 AS `Cantidad`,'' AS `Tipo de Recargo y Descuento`,`f`.`total` AS `Precio Unitario`,'' AS `Descuento3`,'N' AS `Tipo de Descuento`,'VENTAS' AS `Tipo de Venta`,`f`.`total` AS `Total del Producto`,`f`.`total` AS `Precio Lista`,`f`.`total` AS `Total Neto`,'' AS `Ficha Producto`,'EMPDNSADM000000' AS `Centro de Negocios Producto`,'' AS `Producto	Clasificador 1`,'' AS `Producto	Clasificador 2`,'' AS `Producto	Cantidad de Unidad Equivalente`,'' AS `Cantidad de Periodos`,concat('Conf: ',`f`.`conf_no`,' - ',`f`.`razon_social`) AS `Comentario Producto`,'' AS `Análisis Atributo1 Producto`,'' AS `Análisis Atributo2 Producto`,'' AS `Análisis Atributo3 Producto`,'' AS `Análisis Atributo4 Producto`,'' AS `Análisis Atributo5 Producto`,'' AS `Análisis Lote Producto`,'' AS `Fecha de Vencimiento Lote`,'' AS `Ingreso Manual`,'' AS `Tipo de Inventario`,'' AS `Clasificador 1 Inventario Línea`,'' AS `Clasificador 2 Inventario Línea`,'' AS `Número de Descuento`,'' AS `Descuento2`,'' AS `Tipo de Descuento2`,1 AS `Numero de Impuesto`,'IVA' AS `Código de Impuesto`,19 AS `Valor de Impuesto`,`f`.`iva` AS `Monto de Impuesto`,'' AS `Centro de Negocios Producto2`,'' AS `Clasificador 1 Impuesto`,'' AS `Clasificador 2 Impuesto`,1 AS `Número de Cuota`,date_format((`f`.`fecha` + interval 30 day),'%d-%m-%Y') AS `fecha cuota`,`f`.`total` AS `Monto de Cuota`,'' AS `Relación linea Series`,'' AS `Sufijo Artículo Inventario`,'' AS `Prefijo Artículo Inventario`,'' AS `Serie Artículo Inventario`,'' AS `Distrito`,1 AS `Transacción`,date_format(`f`.`fecha`,'%d-%m-%Y') AS `Fecha Facturación Desde`,date_format(`f`.`fecha`,'%d-%m-%Y') AS `Fecha Facturación Hasta`,'' AS `Vía de Transporte`,'' AS `País Destino Receptor`,'' AS `País Destino Embarque`,'' AS `Modalidad Venta`,'' AS `Tipo Despacho`,'' AS `Indicador de Servicio`,'' AS `Claúsula de Venta`,'' AS `Total Claúsula de Venta`,'' AS `Puerto Embarque`,'' AS `Puerto Desembarque`,'' AS `Unidad de Medida Tara`,'' AS `Total Medida Tara`,'' AS `Unidad Peso Bruto`,'' AS `Total Peso Bruto`,'' AS `Unidad Peso Neto`,'' AS `Total Peso Neto`,'' AS `Tipo de Bulto`,'' AS `Total de Bultos`,'' AS `Forma de Pago`,'' AS `Tipo Documento Asociado`,'' AS `Folio Documento Asociado`,'' AS `Fecha Documento Asociado`,'' AS `Comentario Documento Asociado`,'' AS `Email`,'S' AS `Es documento de traspaso`,'' AS `Contacto` 
            
            from `dunasdb`.`fiscales` `f` 
            
            where (`f`.`tipo` = 39)
            and `f`.`fecha` BETWEEN '{desde}' AND '{hasta}'
        

                """
        cursor = conexion.cursor()
        cursor.execute(query)
        result = cursor.fetchall()
        nombres_columnas = [columna[0] for columna in cursor.description]

        output = io.BytesIO() # Buffer en memoria

        # Crear un nuevo archivo .xls
        libro_excel = xlwt.Workbook()
        hoja = libro_excel.add_sheet('Resultados')

        # Escribir los nombres de las columnas
        for columna_index, nombre_columna in enumerate(nombres_columnas):
            hoja.write(0, columna_index, nombre_columna)

        """ # Escribir los datos
        for fila_index, fila_datos in enumerate(result):
            for columna_index, valor in enumerate(fila_datos):
                hoja.write(fila_index + 1, columna_index, str(valor)) """
        for fila_index, fila_datos in enumerate(result):
            for columna_index, valor in enumerate(fila_datos):
                if isinstance(valor, (int, float, )):
                    hoja.write(fila_index + 1, columna_index, valor)
                elif isinstance(valor, str):
                    hoja.write(fila_index + 1, columna_index, valor)
                
                elif valor is None:
                    hoja.write(fila_index + 1, columna_index, "")
                
                else:
                    hoja.write(fila_index + 1, columna_index, str(valor))
        

        
        # Guardar el libro en el buffer de memoria
        libro_excel.save(output)
        output.seek(0) # Mover el cursor al inicio del buffer

        print(f"Archivo {filename} generado en memoria.")
        # Devolver el buffer y el nombre del archivo
        return output, filename

    except mysql.connector.Error as err:
        print(f"Error de base de datos durante la exportación: {err}")
        # Podrías lanzar una excepción más específica o devolver None
        raise ValueError(f"Error de base de datos: {err}") from err
    except Exception as e:
        print(f"Ocurrió un error inesperado durante la exportación: {e}")
        raise ValueError(f"Error inesperado: {e}") from e
    finally:
        if conexion and conexion.is_connected():
            conexion.close()
            print("Conexión a la base de datos cerrada (exportación).")

def procesar_fiscal_stream(filepath):
    """
    Procesa el archivo fiscal y genera (yields) mensajes de progreso.
    """
    conexion = None # Inicializa la conexión fuera del try para el finally
    try:
        # Intenta conectar a la base de datos primero
        yield "Conectando a la base de datos..."
        conexion = connect()
        if not conexion or not conexion.is_connected():
            yield "Error: No se pudo conectar a la base de datos."
            return # Termina la ejecución si no hay conexión

        yield "Conexión a la base de datos establecida."
        yield "Abriendo el archivo Excel..."
        workbook = openpyxl.load_workbook(filepath)
        yield f"Archivo {filepath.filename} abierto. Procesando hojas..." # Usar filepath.filename si es un objeto FileStorage

        hojas = workbook.sheetnames
        hoja = workbook.active
        total_rows = hoja.max_row - 2 # Asumiendo que los datos empiezan en la fila 3
        processed_rows = 0

        yield f"Procesando {total_rows} filas..."

        for i, fila in enumerate(hoja.iter_rows(min_row=3), start=1):  # Comenzar desde la fila 3
            valores = [celda.value for celda in fila]

            # Validar si la fila está vacía o incompleta (ejemplo básico)
            if not any(valores) or valores[1] is None or valores[2] is None or valores[6] is None:
                yield f"Advertencia: Saltando fila {i+2} por datos faltantes o vacía."
                continue

            fecha = valores[0]
            tipo = int(valores[1]) # Asegurar que sea int para la API
            folio = int(valores[2]) if valores[2] is not None else 0 # Asegurar que sea int
            anticipo = valores[3] if valores[3] is not None else ''
            conf_no = valores[4]
            razon_social = valores[5] if valores[5] else 'S/Razón Social' # Evitar None
            rut_raw = valores[6]
            rut = formatRut(str(rut_raw)) if rut_raw is not None else '' # Convertir a str antes de formatear
            neto = valores[7] if valores[7] is not None else 0
            exento = valores[8] if valores[8] is not None else 0
            iva = valores[9] if valores[9] is not None else 0
            total = valores[10] if valores[10] is not None else 0
            cajero = valores[11]
            index = f"{tipo}_{folio}"
            xml_data = None
            origen = 'Opera'

            #primero verificaremos si la factura ya se encuentra en nuestra base de datos local
            query = f"""
                SELECT COUNT(*) FROM fiscales WHERE tipo = {tipo} AND folio = {folio}
            """
            cursor = conexion.cursor()
            cursor.execute(query)
            result = cursor.fetchone()
            cursor.close()

            if result[0] > 0:
                yield f"  -> DTE {tipo}-{folio} ya existe en la base de datos. Saltando inserción."
                continue
            else:
                yield f"  -> DTE {tipo}-{folio} no existe en la base de datos. Procediendo a insertar."


                # Mensaje de progreso por fila
                yield f"Procesando fila {i}/{total_rows}: DTE {tipo}-{folio}, RUT {rut}, Total {total}"

                try:
                    # Consultar estado SII (puede tardar)
                    yield f"  -> Consultando estado SII para {tipo}-{folio}..."
                    # Usar el rutEmisor definido en gde.py o pasarlo como argumento si varía
                    estadoSii_response = consultaEstado(tipo, folio, rutEmisor)
                    estadoSii = estadoSii_response.get('Data', 'Error en respuesta API') # Obtener Data o mensaje de error
                    yield f"  -> Estado SII: {estadoSii}"

                    # Consultar PDF (puede tardar)
                    yield f"  -> Consultando PDF para {tipo}-{folio}..."
                    pdf_response = consultaPDF(tipo, folio, rutEmisor)
                    pdf_data = pdf_response.get('Data', 'Error en respuesta API') # Obtener Data o mensaje de error
                    # Considera no incluir el PDF completo en el mensaje de progreso si es muy grande
                    yield f"  -> PDF obtenido (ID/Estado: {pdf_data[:50]}...)" # Mostrar solo una parte

                    # Insertar xml en la base de datos (replica de la funcion para ejecutar por línea de comandos)
                    print( f" {ASK_SYMBOL} -> Consultando XML para {tipo}-{folio}...")
                    try:
                        payload_xml = {
                                        "Environment": "P",
                                        "Group": "E",
                                        "Rut": rutEmisor,
                                        "DocType": tipo,
                                        "Folio": folio,
                                        "IsForDistribution": "false"
                                        }
                        xml_data = recuperar_xml(payload_xml)['RecoverDocumentResult']['Data']
                        print( colored(f" {OK_SYMBOL} -> XML obtenido (XML Base64 ecoded: {str(xml_data)[:50]}...)", "green"))
                    except Exception as e:
                        print( colored(f" {ERROR_SYMBOL} -> Error al recuperar XML: {e}", "red"))
                        xml_data = None
                    # Inserción en la base de datos
                    print(fecha)
                    query = f"""
                        INSERT IGNORE INTO fiscales
                        (`fecha`, `tipo`, `folio`, `anticipo`, `conf_no`, `razon_social`, `rut`, `neto`, `exento`, `iva`, `total`, `cajero`, `index`, `estadoSii`, `pdf`, `xml`, `origen`)
                        VALUES
                        (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    # Usar parámetros previene inyección SQL y maneja tipos de datos
                    params = (
                        fecha, tipo, folio, anticipo, conf_no, razon_social, rut,
                        neto, exento, iva, total, cajero, index, estadoSii, pdf_data, xml_data, origen
                    )

                    yield f"  -> Insertando DTE {tipo}-{folio} en la base de datos..."
                    cursor = conexion.cursor()
                    cursor.execute(query, params)
                    conexion.commit()
                    rows_affected = cursor.rowcount
                    cursor.close()
                    if rows_affected > 0:
                        yield f"  -> DTE {tipo}-{folio} insertado correctamente."
                    else:
                        yield f"  -> DTE {tipo}-{folio} ya existía (o error al insertar)."
                    processed_rows += 1

                except requests.exceptions.RequestException as api_err:
                    yield f"  -> Error de API consultando {tipo}-{folio}: {api_err}"
                except mysql.connector.Error as db_err:
                    yield f"  -> Error de Base de Datos insertando {tipo}-{folio}: {db_err}"
                    conexion.rollback() # Revertir si falla la inserción
                except Exception as e:
                    yield f"  -> Error inesperado procesando fila {i+2} (DTE {tipo}-{folio}): {e}"

                # Pequeña pausa opcional para que el navegador pueda renderizar las actualizaciones
                # time.sleep(0.05)

        workbook.close()
        yield f"Procesamiento completado. {processed_rows} de {total_rows} filas procesadas e intentadas insertar."

    except FileNotFoundError:
        yield f"Error: Archivo no encontrado."
    except openpyxl.utils.exceptions.InvalidFileException:
        yield "Error: El archivo subido no parece ser un archivo Excel válido (.xlsx)."
    except mysql.connector.Error as err:
        yield f"Error de conexión inicial a la base de datos: {err}"
    except Exception as e:
        yield f"Error general durante el procesamiento: {e}"
    finally:
        if conexion and conexion.is_connected():
            conexion.close()
            yield "Conexión a la base de datos cerrada."

def subir_a_defontana_por_fechas_stream(desde, hasta, tipoDte): #version de erwinlh
    
    fecha_desde = convertir_fecha_iso_a_fecha_hora(desde)
    fecha_hasta = convertir_fecha_iso_a_fecha_hora(hasta)
    
    query_folios_a_insertar = f"select tipo, folio, fecha from fiscales where fecha between '{fecha_desde}' and '{fecha_hasta}' and tipo = {tipoDte} and estadodefontana is null order by idfiscales asc;"
    print(query_folios_a_insertar)
        
    #print(query_folios_a_insertar)
    listado_folios_a_insertar = ejecutar_consulta(connect(), query_folios_a_insertar)
    if len(listado_folios_a_insertar) == 0:
        yield f"{OK_SYMBOL} Ya se encuentran todos los folios en esas fechas"

    #print(listado_folios_a_insertar)
    for lines in listado_folios_a_insertar:
        #print(lines)
        if lines[0] == 33:
            tipodteforDefontana = "FEAV"
        elif lines[0] == 39:
            tipodteforDefontana = "BEAV"    
        


        payload_dte = { 
            "Environment" : "P",            # Se usará T para la Homologación y P para Producción.
            "Group" : "E",                  # Se usará E para Emitidos y R para Recibidos
            "Rut" : "76708884-1", 
            "DocType" : lines[0],               # 33=Factura, 39=Boleta, 41=Nota de Crédito, 52=Nota de Débito, 61= Nota de Crédito Electrónica, 56=Factura Electrónica, 61=Factura Exenta Electrónica, 52=Factura Exenta Electrónica, 110=Factura Exportación Electrónica
            "Folio" : lines[1], 
            "IsForDistribution" : False 
            } 
        
        #print(dte)

        
        try:
            
            print(colored(f"{ASK_SYMBOL} -> Consultando XML para DTE tipo {lines[0]} con folio {lines[1]} a defontana...","white"))
            xml_decodificado = decodificar_xml(recuperar_xml(payload_dte)) ## error al recuperar el XML decodificado
            #print(xml_decodificado)
                
                #print(xml_decodificado)
            try:
                
                payload_cliente = diccionario_a_payload_cliente(xml_decodificado)
                #print(payload_cliente)
            except:
                print(colored(f"\n\n {ERROR_SYMBOL}Error al crear el payload del cliente para DTE tipo {lines[0]} con folio {lines[1]}, de fecha {convertir_datetime_a_fecha_iso(lines[2])}", "red"))
                yield "\n\n {ERROR_SYMBOL}Error al crear el payload del cliente para DTE tipo {lines[0]} con folio {lines[1]}"
            
            yield f"\n{ASK_SYMBOL}Consultando venta tipo {lines[0]} con folio {lines[1]}..."
            consultar_ventas = consultar_venta(tipodteforDefontana, lines[1])
            
                            
            if consultar_ventas == 200:
                print(colored(f" {WARNING_SYMBOL}DTE tipo {lines[0]} con folio {lines[1]} se encuentra contabilizado en Defontana, continue...","yellow"))
                #yield(colored(f" {WARNING_SYMBOL}DTE tipo {lines[0]} con folio {lines[1]} se encuentra contabilizado en Defontana, continue...","yellow"))
                yield f"  {WARNING_SYMBOL}DTE tipo {lines[0]} con folio {lines[1]} se encuentra contabilizado en Defontana, continue..."
                try:
                    update_query = f"update fiscales set estadoDefontana='CENTRALIZADO' where tipo = {lines[0]} and folio = {lines[1]};"
                    
                    #marca el dte como CENTRALIZADO porque ya se encuentra contabilizado en defontana
                    insertar_resultados(connect(), update_query)
                    print(colored(f"  {OK_SYMBOL} DTE tipo {lines[0]} con folio {lines[1]} actualizado en DB Local con estado CENTRALIZADO","green"))
                    yield f"  {OK_SYMBOL} DTE tipo {lines[0]} con folio {lines[1]} actualizado en DB Local con estado CENTRALIZADO"
                except Exception as e:
                    print(colored(f"  {ERROR_SYMBOL} Error al actualizar DTE tipo {lines[0]} con folio {lines[1]}: {e}","red"))
                    yield f"  {ERROR_SYMBOL} Error al actualizar DTE tipo {lines[0]} con folio {lines[1]}: {e}"
                    
            else:
                #resto de la lófica, primero verificar si cliente existe, si no crearlo, luego insertar el documento
                
                if consultar_cliente(payload_cliente["legalCode"])['message'] == 'La busqueda no produjo ningún resultado':
                    #print(f"\nCliente {payload_cliente['legalCode']} no existe en Defontana, se procederá a crear el cliente")
                    try:
                        print(colored(f" {NEW_SYMBOL} Cliente nuevo, guardando cliente {payload_cliente['legalCode'], payload_cliente['name'], payload_cliente['address'], payload_cliente['district'], payload_cliente['email'], payload_cliente['city']}", "blue")) 
                        yield f" {NEW_SYMBOL} Cliente nuevo, guardando cliente {payload_cliente['legalCode'], payload_cliente['name'], payload_cliente['address'], payload_cliente['district'], payload_cliente['email'], payload_cliente['city']}"
                        guardar_cliente(payload_cliente) 
                    except Exception as e:
                        print(f"Error al guardar cliente {payload_cliente['legalCode']}: {e}")
                        yield f"Error al guardar cliente {payload_cliente['legalCode']}: {e}"
                        #print(xml_decodificado)
                        
                        
                else:
                    #print(colored(f" {OK_SYMBOL,'green'} Cliente {payload_cliente['legalCode']}, {payload_cliente['name']}: ya existe en Defontana...","green"))
                    print(colored(f" {OK_SYMBOL} Cliente {payload_cliente['legalCode']}, {payload_cliente['name']}: ya existe en Defontana...","green"))
                    yield f" {OK_SYMBOL} Cliente {payload_cliente['legalCode']}, {payload_cliente['name']}: ya existe en Defontana..."
                

                payload_factura = diccionario_a_payload_factura(xml_decodificado)
                #print(payload_factura)
                
                #print(f"\nPayload del DTE tipo {lines[0]} con folio {lines[1]}:\n{payload_factura}")

                try:
                    print(f" {colored(NEW_SYMBOL,'blue')} Guardando DTE tipo {lines[0]} con folio {lines[1]} en Defontana...")
                    #yield (f" {colored(NEW_SYMBOL,'blue')} Guardando DTE tipo {lines[0]} con folio {lines[1]} en Defontana...")
                    yield f" {NEW_SYMBOL} Guardando DTE tipo {lines[0]} con folio {lines[1]} en Defontana..."
                    
                    guardar_venta_result = guardar_venta(payload_factura)
                    
                    #print(guardar_venta_result)
                    if guardar_venta_result.status_code == 200:
                        update_query = f"update fiscales set estadoDefontana='CENTRALIZADO' where tipo = {lines[0]} and folio = {lines[1]};"                                          
                        insertar_resultados(connect(), update_query)
                        print(colored(f"  {OK_SYMBOL} DTE tipo {lines[0]} con folio {lines[1]} guardado en Defontana y actualizado en DB Local con estado CENTRALIZADO","green"))
                        yield f"  {OK_SYMBOL} DTE tipo {lines[0]} con folio {lines[1]} guardado en Defontana y actualizado en DB Local con estado CENTRALIZADO"
                    else:
                        print(guardar_venta_result.text)
                except Exception as e:
                    print(f"Error al guardar DTE tipo {lines[0]} con folio {lines[1]}: {e}")
                    yield f"Error al guardar DTE tipo {lines[0]} con folio {lines[1]}: {e}"
                    #print(payload_factura)
                    
        except Exception as e:
            #print(colored(f"{ERROR_SYMBOL}Error al decodificar XML para DTE tipo {lines[0]} con folio {lines[1]}, de fecha {lines[2]}: {e}","red"))        
            yield f"error al decodificar XML para DTE tipo {lines[0]} con folio {lines[1]}, de fecha {convertir_datetime_a_fecha_iso(lines[2])}: {e}"

def procesar_fiscal_gde_stream(filepath):
    """
    Procesa un archivo fiscal CSV (tabulado) y genera (yields) mensajes de progreso.
    """
    conexion = None # Inicializa la conexión fuera del try para el finally
    try:
        # Intenta conectar a la base de datos primero
        yield "Conectando a la base de datos..."
        conexion = connect()
        if not conexion or not conexion.is_connected():
            yield "Error: No se pudo conectar a la base de datos."
            return # Termina la ejecución si no hay conexión
        
        yield "Abriendo archivo CSV..."
        df = pd.read_csv(filepath, sep="\t", encoding='latin1')  # usa tabulador como separador
        total_rows = len(df)
        yield f"Archivo abierto. {total_rows} filas encontradas."

        for idx, row in df.iterrows():
            try:
                # Validar campos requeridos
                if pd.isnull(row["Tipo DTE"]) or pd.isnull(row["Folio"]) or pd.isnull(row["RUT Cliente"]):
                    yield f"Advertencia: Saltando fila {idx+2} por datos faltantes obligatorios."
                    continue

                fecha = row["Emisión"]
                tipo = int(row["Tipo DTE"])
                folio = int(row["Folio"]) if pd.notnull(row["Folio"]) else 0
                anticipo = ''
                conf_no = ''
                razon_social = row["Cliente"] if pd.notnull(row["Cliente"]) else 'S/Razón Social'
                rut = formatRut(str(row["RUT Cliente"])) if pd.notnull(row["RUT Cliente"]) else ''
                neto = row["Monto Neto"] if pd.notnull(row["Monto Neto"]) else 0
                exento = row["Monto Exento"] if pd.notnull(row["Monto Exento"]) else 0
                iva = row["IVA"] if pd.notnull(row["IVA"]) else 0
                total = row["Monto Total"] if pd.notnull(row["Monto Total"]) else 0
                cajero = row["Nombre POS"] if pd.notnull(row["Nombre POS"]) else ''
                index = f"{tipo}_{folio}"

                yield f"Procesando fila {idx+1}/{total_rows}: DTE {tipo}-{folio}, RUT {rut}, Total {total}"

                # Consultar estado SII
                yield f"  -> Consultando estado SII para {tipo}-{folio}..."
                estadoSii_response = consultaEstado(tipo, folio, rutEmisor)
                estadoSii = estadoSii_response.get('Data', 'Error en respuesta API')
                yield f"  -> Estado SII: {estadoSii}"

                # Consultar PDF
                yield f"  -> Consultando PDF para {tipo}-{folio}..."
                pdf_response = consultaPDF(tipo, folio, rutEmisor)
                pdf_data = pdf_response.get('Data', 'Error en respuesta API')
                yield f"  -> PDF obtenido (ID/Estado: {str(pdf_data)[:50]}...)"

                # Consultar XML
                payload_xml = {
                    "Environment": "P",
                    "Group": "E",
                    "Rut": rutEmisor,
                    "DocType": tipo,
                    "Folio": folio,
                    "IsForDistribution": "false"
                    }
                
                yield f"  -> Consultando XML para {tipo}-{folio}..."
                xml_response = recuperar_xml(payload_xml)
                xml_data = xml_response.get('Data', 'Error en respuesta API')
                yield f"  -> XML obtenido (ID/Estado: {str(pdf_data)[:50]}...)"                
                
                query = f"""
                    INSERT IGNORE INTO fiscales
                    (`fecha`, `tipo`, `folio`, `anticipo`, `conf_no`, `razon_social`, `rut`, `neto`, `exento`, `iva`, `total`, `cajero`, `index`, `estadoSii`, `pdf`, `xml`)
                    VALUES
                    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                # Usar parámetros previene inyección SQL y maneja tipos de datos
                params = (
                    fecha, tipo, folio, anticipo, conf_no, razon_social, rut,
                    neto, exento, iva, total, cajero, index, estadoSii, pdf_data, xml_data
                )

                print(params)
                input("Presione una tecla para continuar...")

                yield f"  -> Insertando DTE {tipo}-{folio} en la base de datos..."
                cursor = conexion.cursor()
                cursor.execute(query, params)
                conexion.commit()
                rows_affected = cursor.rowcount
                cursor.close()
                if rows_affected > 0:
                    yield f"  -> DTE {tipo}-{folio} insertado correctamente."
                else:
                    yield f"  -> DTE {tipo}-{folio} ya existía (o error al insertar)."
                processed_rows += 1

            except requests.exceptions.RequestException as api_err:
                yield f"  -> Error de API consultando {tipo}-{folio}: {api_err}"
            except Exception as e:
                yield f"  -> Error inesperado procesando fila {idx+2} (DTE {tipo}-{folio}): {e}"

        yield f"Procesamiento completado. {total_rows} filas leídas del archivo."

    except FileNotFoundError:
        yield f"Error: Archivo no encontrado."
    except mysql.connector.Error as err:
        yield f"Error de conexión inicial a la base de datos: {err}"

    except pd.errors.ParserError:
        yield "Error: El archivo no se pudo leer correctamente. ¿Está separado por tabulaciones?"
    except Exception as e:
        yield f"Error general durante el procesamiento: {e}"