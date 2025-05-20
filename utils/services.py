import pandas as pd
import openpyxl
import mysql.connector
import requests
import json
import os
from dotenv import load_dotenv
from termcolor import colored
from datetime import datetime
from datetime import timedelta
from utils.common import formatRut, convertir_fecha_latin_a_iso, formatear_rut, ejecutar_consulta, insertar_resultados, connect, convertir_datetime_a_fecha_iso_hora

#entorno gde
load_dotenv()
apiKey =os.getenv("apiKeyGde")
apiServergde = os.getenv("apiServerGde")
apiEnviroment = 'P'
apiGroup = 'E'

#define si apunta a produccion o homologacion
apiValue = 'api' #para produccion debe ser 'api', para homologacion 'replapi'


rutEmisor = '76708884-1'
tipoDte = 33
folioDte = 310

# Define symbols
ERROR_SYMBOL = " ✗ "
OK_SYMBOL = " ✓ "
ASK_SYMBOL = " ? "
NEW_SYMBOL = " # "
WARNING_SYMBOL = " ! "


#funciones de GDExpress
def consultaEstado(tipo, folio, rutEmisor):    
    tipoDte = tipo
    folioDte = folio

    apiUrl = f'http://{apiServergde}/api/Core.svc/core/FiscalStatus/{apiEnviroment}/{apiGroup}/{rutEmisor}/{tipoDte}/{folioDte}'

    headers = {'AuthKey': apiKey,
                'Content-Type': "application/json",
                "Accept": "application/json"}

    r = requests.get(apiUrl, headers=headers)
    data = r.json()
    return data

def actualizar_estado_sii_masivo():
    query = f"select tipo, folio from fiscales where estadosii is NULL"

    listado = ejecutar_consulta(connect(), query)
    registros = len(listado)
    if registros == 0:
        print(colored(f"{OK_SYMBOL} No se encontraron registros para actualizar el estadosii", "green"))
    else:
        print(colored(f"{NEW_SYMBOL} Se encontraron {registros} registros para actualizar el estadosii", "blue"))
    for lines in listado:
        #recuperar el estadosii
        estadosii = consultaEstado(lines[0], lines[1],rutEmisor)
        estadosii = estadosii['Data']

        #actualizar el estadosii en la db local
        update_query = f"update fiscales set estadosii = '{estadosii}' where tipo = {lines[0]} and folio = {lines[1]}"
        
        if estadosii == 'Aprobado':
            print(colored(f" {NEW_SYMBOL} Actualizando el estadosii del folio {lines[0]}-{lines[1]} con el estado: {estadosii}","blue"))
        else:
            print(colored(f" {WARNING_SYMBOL} Actualizando el estadosii del folio {lines[0]}-{lines[1]} con el estado: {estadosii}","yellow"))
            #alert por estado distinto a aprobado
        try:

            
            insertar_resultados(connect(), update_query)
            registros -=1
            print(colored(f" {OK_SYMBOL} Quedan {registros} registros por actualizar el estadosii", "white"))
        except:
            print(f"Error al actualizar el estadosii del folio {lines[1]} con el estado {estadosii}")
            break

def actualizar_xml_masivo():
    query = f"select tipo, folio from fiscales where xml is NULL"

    listado = ejecutar_consulta(connect(), query)
    registros = len(listado)
    print(colored(f"  {NEW_SYMBOL} Se encontraron {registros} registros para actualizar el XML", "blue"))
    for lines in listado:
        print(lines)
        #recuperar el estadosii
        dte = { 
            "Environment" : "P",            # Se usará T para la Homologación y P para Producción.
            "Group" : "E",                  # Se usará E para Emitidos y R para Recibidos
            "Rut" : "76708884-1", 
            "DocType" : lines[0],               # 33=Factura, 39=Boleta, 41=Nota de Crédito, 52=Nota de Débito, 61= Nota de Crédito Electrónica, 56=Factura Electrónica, 61=Factura Exenta Electrónica, 52=Factura Exenta Electrónica, 110=Factura Exportación Electrónica
            "Folio" : lines[1], 
            "IsForDistribution" : "false" 
            }
        print(dte)
        print(colored(f"  {NEW_SYMBOL} Consultando el XML del folio {lines[0]}-{lines[1]}...", "blue"))
        try:
            xml_recuperado = recuperar_xml(dte, apiServergde)
            xml = xml_recuperado['RecoverDocumentResult']['Data']
            print(xml[:100])
            
        except:
            print(colored(f"  {ERROR_SYMBOL} Error al recuperar el XML del folio {lines[0]}-{lines[1]}", "red"))
            #print(xml)
        
        #actualizar el xml en la db local
        update_query = f"update fiscales set xml = '{xml}' where tipo = {lines[0]} and folio = {lines[1]}"
        print(colored(f" {NEW_SYMBOL}Actualizando el xml del folio {lines[0]}-{lines[1]} con el estado {xml[:50]}","white"))
        try:
            
            #insertar_resultados(connect(), update_query)
            registros -=1
            print(colored(f" {OK_SYMBOL} Quedan {registros} registros por actualizar el XML", "yellow"))

        except:
            #print(f"Error al actualizar el xml del folio {lines[0]} con el estado {xml[:50]}")
            print(colored(f"  {ERROR_SYMBOL} Error al actualizar el xml del folio {lines[0]}-{lines[1]} con el estado {xml[:50]}", "red"))
            
def consultaPDF(tipo, folio, rutEmisor):
    
    tipoDte = tipo
    folioDte = folio

    apiUrl = f'http://{apiServergde}/api/Core.svc/core/RecoverPDF_V2/{apiEnviroment}/{apiGroup}/{rutEmisor}/{tipoDte}/{folioDte}'

    headers = {'AuthKey': apiKey,
                'Content-Type': "application/json",
                "Accept": "application/json"}

    r = requests.get(apiUrl, headers=headers)
    data = r.json()
    return data

def salvarDataPdf(data): #funcion que se hizo para los dte que ya existian en la base de datos, POC
    pdfData = data
    idfiscalpdf = pdfData["Result"]
    dataPdf = pdfData["Data"]

    queryFiscalPDF = f"INSERT INTO dunasdb.fiscalpdf (`rutEmisor`,`tipoDte`,`folioDte`,`dataPdf`) VALUES ( \'{rutEmisor}\', \'{tipoDte}\', \'{folioDte}\', \'{dataPdf}\')"
    
    conexion = connect()
    insertar_resultados(conexion, queryFiscalPDF)



def recuperar_xml(payload): #devuelve un string en base 64
    """
    Realiza un POST a http://{apiserver}/api/Core.svc/core/RecoverXML_V2
    con el payload solicitado y convierte la respuesta XML a un diccionario.

    Args:
        payload (dict): Diccionario con los campos requeridos por la API.
        apiserver (str): Hostname/IP y puerto del API server.

    Returns:
        dict: Respuesta de la API convertida de XML a diccionario.

    Ejemplo de uso:
        payload = {
            "Environment": "P",
            "Group": "E",
            "Rut": "76129486-5",
            "DocType": "33",
            "Folio": "20",
            "IsForDistribution": "false"
        }
        respuesta = recuperar_xml(payload, "apitest.ejemplo.cl:1234")
    """
    # Se requiere xmltodict: pip install xmltodict
    try:
        import xmltodict
    except ImportError:
        raise ImportError("Debes instalar xmltodict: pip install xmltodict")

    apiUrl = f'http://{apiServergde}/api/Core.svc/core/RecoverXML_V2'
    headers = {
        'AuthKey': apiKey,
        'Content-Type': "application/json",
        "Accept": "application/xml"
    }
    try:
        r = requests.post(apiUrl, headers=headers, json=payload)
        #print(r)
        if r.status_code == 200:
            try:
                xml_data = r.text
                dict_data = xmltodict.parse(xml_data)
                return dict_data
            except Exception as ex:
                print("Error parseando XML:", ex)
                return None
        else:
            print(f"Error en request: {r.status_code} {r.text}")
            return None
    except Exception as e:
        print(f"Excepción en recuperar_xml: {e}")
        return None

#funciones de Defontana

# se usa cuando se necesita, para conseguir un nuevo token
def get_token():
    """
    Obtiene y almacena el access_token para la API de Defontana.
    - Si ya existe un token almacenado, pregunta al usuario si desea validarlo o renovarlo.
    - Si no existe o se decide regenerar, realiza la autenticación.
    El token se almacena/lee desde token.conf (formato JSON) en el mismo directorio.
    """
    token_file = "token.conf"
    token_data = None

    # Verificar si hay un token ya guardado
    if os.path.isfile(token_file):
        with open(token_file, "r", encoding="utf-8") as f:
            try:
                token_data = json.load(f)
                access_token = token_data.get("access_token")
                print("Token encontrado en token.conf.")
                resp = input("¿Deseas ver el token existente (V) o regenerar uno nuevo (N)? [V/N]: ").strip().upper()
                if resp == "V":
                    # Placeholder para la función futura de validación
                    print("Función checkToken() aún no implementada.")
                    return os.system("type token.conf")# Aquí irá la llamada: checkToken(access_token)
                    # Por ahora se retorna el token existente
                    return access_token
                elif resp == "N":
                    print("Se generará un nuevo token...")
                else:
                    print("Opción no reconocida. Usando token existente por defecto.")
                    return access_token
            except Exception as e:
                print("Error leyendo token.conf, se solicitará uno nuevo:", e)
    # Si no hay token o hay que renovarlo, realizar autenticación
    client = os.getenv("client")
    company = os.getenv("company")  #este valor debe cambiar para cada empresa
    user = os.getenv("user")
    password = os.getenv("password")

    # Construir la URL de autenticación y llamar a la API
    base_auth_url = f"https://{apiValue}.defontana.com/api/Auth"
    params = {
        "client": client,
        "company": company,
        "user": user,
        "password": password
    }
    try:
        resp = requests.get(base_auth_url, params=params)
        if resp.status_code == 200:
            response_json = resp.json()
            access_token = response_json.get("access_token") or response_json.get("accessToken")
            if access_token:
                nueva_data = {"access_token": access_token}
                with open(token_file, "w", encoding="utf-8") as f:
                    json.dump(nueva_data, f, indent=4)
                print("Nuevo access_token almacenado en token.conf.")
                return access_token
            else:
                print("No se encontró 'access_token' en la respuesta:", response_json)
        else:
            print(f"Error autenticando: {resp.status_code} {resp.text}")
    except Exception as e:
        print(f"Error al hacer la request de autenticación: {e}")
    return None

# Función para obtener el header de autorización utilizando el token actual almacenado en token.conf
def get_auth_headers():
    #token_file = "token.conf"
    access_token = os.getenv("token")
    if not access_token:
        print("No hay token disponible. Ejecuta getToken() primero.")
    return {
        "Authorization": f"Bearer {os.getenv('token')}"
    } if access_token else {}



def consultar_empresa():
    """
    Recupera y muestra los datos de la empresa utilizando el token almacenado en token.conf.
    """
    url_company = f"https://{apiValue}.defontana.com/api/Company"
    headers = get_auth_headers()
    if not headers:
        print("No se encontró un token válido. Usa primero getToken() para autenticarte.")
        return None
    try:
        resp = requests.get(url_company, headers=headers)
        if resp.status_code == 200:
            print("Respuesta de la API (Company):")
            print(resp.text)
            # Si prefieres, también puedes retornar resp.text para usarlo luego en el código
            return resp.text
        else:
            print(f"Error solicitando información de empresa: {resp.status_code} {resp.text}")
            return None
    except Exception as e:
        print(f"Error en get_company_info(): {e}")
        return None

def consultar_cliente(legalCode):
    """
    Consulta uno o más clientes en Defontana usando el RUT (legalCode) con el endpoint correcto.
    Args:
        legalCode (str): RUT del cliente a consultar (se formatea automáticamente).
        status (int): Estado (default 0).
        itemsPerPage (int): Cantidad de items por página (default 10).
        pageNumber (int): Número de página (default 1).

    Returns:
        dict o str: Respuesta de la API.

    Ejemplo de uso:
        respuesta = consultar_cliente("16143583-K")
    """
    status=0
    itemsPerPage=10
    pageNumber=1

    url = f"https://{apiValue}.defontana.com/api/Sale/GetClients"
    headers = get_auth_headers()
    if not headers:
        print("No se pudo obtener un token válido para autenticar la petición.")
        return None
    # Formatea el RUT al formato estándar
    legalCode_fmt = formatear_rut(legalCode)
    params = {
        "legalCode": legalCode_fmt,
        "status": status,
        "itemsPerPage": itemsPerPage,
        "pageNumber": pageNumber
    }
    try:
        resp = requests.get(url, headers=headers, params=params)
        if resp.status_code == 200:
            #print("Respuesta GetClients:")
            data = resp.json()
            #print(data)
            return data
        else:
            print(f"Error al consultar cliente: {resp.status_code} {resp.text}")
            return None
    except Exception as e:
        print(f"Excepción en consultar_cliente(): {e}")
        return None
        print(f"Excepción en consultar_cliente(): {e}")
        return None

def diccionario_a_payload_cliente(xml_data):
    
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '33':
            tipoDocumento = "FEAV"
            envioTipo = "EnvioDTE"
            accountNumber = "1110401021"
    except:
        pass
    try:
        if xml_data['EnvioBOLETA']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '39':
            tipoDocumento = "BEAV"
            envioTipo = "EnvioBOLETA"
            accountNumber = "1110401023"
    except:
        pass 
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '34':
            tipoDocumento = "FEEV"
            envioTipo = "EnvioDTE"
    except:
        pass
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '61':
            tipoDocumento = "NCAV"
            envioTipo = "EnvioDTE"
    except:
        pass
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '56':
            tipoDocumento = "NDAV"
            envioTipo = "EnvioDTE"
    except:
        pass 

    try:
        giro = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Emisor']['GiroEmis']
    except:
        giro = 'noGiro'

    try:
        name = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['RznSocRecep']
    except:
        name = 'noName'

    try:
        address = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['DirRecep']
    except:
        address = 'noAddress'

    try:
        comuna = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['CmnaRecep']
    except:
        comuna = 'noComuna'

    try:
        if xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['Contacto'] == None:
            email = 'no@tiene.com'
        else:
            email = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['Contacto']
    except:
        email = 'no@tiene.com'
    
    try:
        ciudad = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['CiudadRecep']
    except:
        ciudad = 'noCiudad'

    try:
        rut = formatear_rut(xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['RUTRecep'])
    except:
        rut = '66.666.666-6'

    payload = client_payload = {
                "legalCode": rut,
                "fileid": rut,
                "name": name,
                "address": address,
                "district": comuna,
                "email": email,
                #"business": "...",
                "rubroId": "1",
                "giro": giro,
                "city": ciudad
                #"customFields": [ """{"numClase": 1, "descC": "ejemplo"}""" ]
            }
    return payload

def guardar_cliente(datos_cliente):
    """
    Envía una solicitud para guardar un cliente en Defontana usando un diccionario de datos.
    Args:
        datos_cliente (dict): Diccionario con los campos requeridos por la API, por ejemplo:
            {
                "legalCode": "...",
                "fileid": "...",
                "name": "...",
                "address": "...",
                "district": "...",
                "email": "...",
                "business": "...",
                "rubroId": "...",
                "giro": "...",
                "city": "...",
                "customFields": [ {"numClase": 1, "descC": "ejemplo"} ]
            }

    Returns:
        dict o str: Respuesta de la API.

    Ejemplo de uso:
        datos_cliente = {
            "legalCode": "CL123",
            "fileid": "123",
            "name": "Juan Pérez",
            "address": "Calle Falsa 123",
            "district": "Centro",
            "email": "juan@email.com",
            "business": "Comercio",
            "rubroId": "1",
            "giro": "Ventas",
            "city": "Santiago",
            "customFields": [ {"numClase": 1, "descC": "Preferente"} ]
        }
        respuesta = guardar_cliente(datos_cliente)
    """
    url = f"https://{apiValue}.defontana.com/api/Sale/SaveClient"
    headers = get_auth_headers()
    if not headers:
        print("No se pudo obtener un token válido para autenticar la petición.")
        return None
    try:
        resp = requests.post(url, headers={**headers, "Content-Type": "application/json"}, json=datos_cliente)
        if resp.status_code in (200, 201):
            print(colored(f"  {OK_SYMBOL} Cliente guardado correctamente.", "green"))
            #print("Cliente guardado correctamente.")
            try:
                data = resp.json()
                #print(data)
                return data
            except Exception:
                print(resp.text)
                return resp.text
        else:
            #print(f"Error al guardar cliente: {resp.status_code} {resp.text}")
            print(colored(f"  {ERROR_SYMBOL} Error al guardar cliente: {resp.status_code} {resp.text}", "red"))
            return None
    except Exception as e:
        print(f"Excepción en guardar_cliente(): {e}")
        return None

def consultar_venta(tipo, folio):
    

    url = f"https://{apiValue}.defontana.com/api/Sale/GetSale"
    headers = get_auth_headers()
    if not headers:
        print("No se pudo obtener un token válido para autenticar la petición.")
        return None
    
    tipodte = tipo
    foliodte = folio
    
    
    params = {
        "documentType": tipodte,
        "number": foliodte
    }


    try:
        print(colored(f"\n{NEW_SYMBOL}Consultando venta tipo {tipodte} con folio {foliodte}...", "blue"))
        resp = requests.get(url, headers=headers, params=params)
        #print(resp)
        if resp.status_code == 200:
            
            
            return resp.status_code
        else:
            #print(f"Error al consultar venta: {resp.status_code} {resp.text}")
            
            return resp.status_code
            #return None
    except Exception as e:
        print(f"Excepción en consultar_venta(): {e} {resp.json()}")
        return None

def diccionario_a_payload_factura(xml_data): 
    """
    
    Convierte la data que viene del XML en los campos para completar el payload
    
    Keyword arguments:
    argument -- description
    Return: return_description
    
    Defontana - Listado de Tipos de Documentos
    BEAV 	39 Boleta Electronica Venta 	                Vta_BEAV	39 - Boleta Electrónica	Activo	
    FEAV 	33 Factura electrónica afecta Venta 	        Vta_FEAV	33 - Factura Electrónica	Activo	
    FEEV 	34 Factura Electrónica Exenta Venta 	        Vta_FEEV	34 - Factura No Afecta o Exenta Electronica	Activo	
    NCAV 	61 Nota de Credito Electronica Venta 	        Vta_NCAV	61 - Nota de Crédito Electrónica	Activo	
    NCEV 	61 Nota de Credito Exenta Electronica Venta 	Vta_NCEV	61 - Nota de Crédito Electrónica	Activo	
    NDAV 	56 Nota de Debito Electronica Venta 	        Vta_NDAV	56 - Nota de Débito Electrónica	Activo	
    NDEV 	56 Nota de Debito Exenta Electronica Venta 	    Vta_NDEV	56 - Nota de Débito Electrónica	Activo 
    
    
    Cuentas Contables 
    1110401021 - Facturas x Cobrar
    1110401022 - Faceturas Exportacion x Cobrar
    1110401023 - Boletas x Cobrar
    
    ejemplo de xml_data:
    {'EnvioDTE': {'@version': '1.0', '@xmlns': 'http://www.sii.cl/SiiDte', '@xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance', '@xsi:schemaLocation': 'http://www.sii.cl/SiiDte EnvioDTE_v10.xsd', 'SetDTE': {'@ID': 'ID17424dbf30f0438eac7e555dded90038', 'Caratula': {'@version': '1.0', 'RutEmisor': '76708884-1', 'RutEnvia': '10974377-1', 'RutReceptor': '76923783-6', 'FchResol': '2024-09-24', 'NroResol': '0', 'TmstFirmaEnv': '2024-09-24T15:15:20', 'SubTotDTE': {'TpoDTE': '33', 'NroDTE': '1'}}, 'DTE': {'@version': '1.0', 'Documento': {'@ID': 'F10T33', 'Encabezado': {'IdDoc': {'TipoDTE': '33', 'Folio': '10', 'FchEmis': '2024-09-24', 'FmaPago': '1'}, 'Emisor': {'RUTEmisor': '76708884-1', 'RznSoc': 'Sociedad Administradora Arata Schiappasse SPA', 'GiroEmis': 'Hoteleria', 'Acteco': '551001', 'Sucursal': 'Marina Dunas', 'DirOrigen': 'Las Tortolas 230, Reñaca, 2542570', 'CmnaOrigen': 'Valparaiso', 'CiudadOrigen': 'Viña del Mar'}, 'Receptor': {'RUTRecep': '76923783-6', 'RznSocRecep': 'AIRBNB', 'GiroRecep': 'Agencia', 'DirRecep': 'Av. Andres Bello 2687', 'CmnaRecep': 'Las Condes', 'CiudadRecep': 'Santiago'}, 'Totales': {'MntNeto': '220000', 'TasaIVA': '19', 'IVA': '41800', 'MntTotal': '261800'}}, 'Detalle': [{'NroLinDet': '1', 'NmbItem': 'Alojamiento(2)', 'QtyItem': '1', 'PrcItem': '200000', 'MontoItem': '200000'}, {'NroLinDet': '2', 'NmbItem': 'Otros(1)', 'QtyItem': '1', 'PrcItem': '20000', 'MontoItem': '20000'}], 'TED': {'@version': '1.0', 'DD': {'RE': '76708884-1', 'TD': '33', 'F': '10', 'FE': '2024-09-24', 'RR': '76923783-6', 'RSR': 'AIRBNB', 'MNT': '261800', 'IT1': 'Alojamiento(2)', 'CAF': {'@version': '1.0', 'DA': {'RE': '76708884-1', 'RS': 'SOCIEDAD ADMINISTRADORA ARATA SCHIAPPACA', 'TD': '33', 'RNG': {'D': '1', 'H': '20'}, 'FA': '2024-07-26', 'RSAPK': {'M': '91/tGvfC4jYK3TcGvfAle9lFn28xuOxS+xF33x2FIfWOCpjAu6Q6CLalCTUZVEEpGpPLC4LAS11+AoXlu/yUgQ==', 'E': 'Aw=='}, 'IDK': '300'}, 'FRMA': {'@algoritmo': 'SHA1withRSA', '#text': 'Oed9tlGWqEwQ+CpDMZ5KaMub0gKHK3LAVhAouhDZ20hD4BX1/d4DS6zhoWSZA670W0HZDxCl4lzIc/Fn8dbq3Q=='}}, 'TSTED': '2024-09-24T15:15:20'}, 'FRMT': {'@algoritmo': 'SHA1withRSA', '#text': 'M69cIYQaHnvhLctFrBxP1CpOpLqVkgVX1ZYPght3UwOUl0t5tX1w4CCHpJluurDTDFP+5FzI9lVFq22PXmil3g=='}}, 'TmstFirma': '2024-09-24T15:15:20'}, 'Signature': {'@xmlns': 'http://www.w3.org/2000/09/xmldsig#', 'SignedInfo': {'CanonicalizationMethod': {'@Algorithm': 'http://www.w3.org/TR/2001/REC-xml-c14n-20010315'}, 'SignatureMethod': {'@Algorithm': 'http://www.w3.org/2000/09/xmldsig#rsa-sha1'}, 'Reference': {'@URI': '#F10T33', 'Transforms': {'Transform': {'@Algorithm': 'http://www.w3.org/TR/2001/REC-xml-c14n-20010315'}}, 'DigestMethod': {'@Algorithm': 'http://www.w3.org/2000/09/xmldsig#sha1'}, 'DigestValue': 'uP8bZ+vO/WMEtKPvg7wWwx+aQOs='}}, 'SignatureValue': 'eN0uy/eZ1rnH60pOkehrgZk1KGf2RyY01dTPGNswrRNky/q6uoxt4Y9rFZPkq0Vl\nIOveujpheNeq6d+Bv/j8L5ox0wWyciIsS3WBkjFBniTpQWL6mJQJ/JJX3fIrI6Dz\n7XC1nRPDgWfK1Rd9e9eS82hC6anCigjTpsxMxmOwLQIvRgyiXJG2UvULuoJq/BGU\nCj+VlVoH2FuqUiW4TZQ5evudRkcZ7DAFmKSc30nVH9PIWiKdbFKsGDlqXTMZoLo0\nejjWTQNjmfVOZlSzpWwUT5kfmOkOD+LWptoLgVRGhZXWiPgL7CK/DoA7H53zusCi\nsTTsD5y4pCB1lDHn0QYRuA==', 'KeyInfo': {'KeyValue': {'RSAKeyValue': {'Modulus': 'h4W6vcB9BqsdG6EQq6Lub/DGavHgVkmust1R9hmoaux/eQr9zMvgw8H+dwf0WDiw\ncD7xTUB4iN8IPCUfhJL2jv413uPVsKSAwiEMWFerTMYV/hWWIY486daglcaUF8SL\nMtn6ool6F63IbcxGAxifqwKOuFyL5hZll3uupxqRJmdLpz7q9KZgLY0tIEiW3JqY\nuPHEZne/R6VrcPN0QR7MwS4mHX6H+s45UpJNLLA7PiINAyCFyBkfJfMUoa8KsISz\nHfMe+gHahdayfrGCLgJRas1Mbo9gsqJAsd1SD6qURzl9IJuYO3TymurwuhFv/gFN\nkUmSgJALa40pFsveMG/xBw==', 'Exponent': 'AQAB'}}, 'X509Data': {'X509Certificate': 'MIIHsDCCBZigAwIBAgIKGLfJNgABAAmikTANBgkqhkiG9w0BAQsFADCBvzELMAkG\nA1UEBhMCQ0wxHTAbBgNVBAgTFFJlZ2lvbiBNZXRyb3BvbGl0YW5hMREwDwYDVQQH\nEwhTYW50aWFnbzEUMBIGA1UEChMLRS1DRVJUQ0hJTEUxIDAeBgNVBAsTF0F1dG9y\naWRhZCBDZXJ0aWZpY2Fkb3JhMR4wHAYDVQQDExVFLUNFUlRDSElMRSBDQSBGRVMg\nMDIxJjAkBgkqhkiG9w0BCQEWF3NjbGllbnRlQGUtY2VydGNoaWxlLmNsMB4XDTI0\nMDMyNzE0NTgxMVoXDTI3MDMyNzE0NTgxMVowgawxCzAJBgNVBAYTAkNMMRMwEQYD\nVQQIEwpWQUxQQVJBSVNPMRQwEgYDVQQHDAtWQUxQQVJBw41TTzEiMCAGA1UEChMZ\nUkVWSUNFTlRSTyBWQUxQQVJBSVNPIFNQQTEKMAgGA1UECwwBKjEiMCAGA1UEAxMZ\nU0VSR0lPIEFOVE9HTk9MSSBDQVNURUxMSTEeMBwGCSqGSIb3DQEJARYPSU5GT0BS\nRVZJVEVDLkNMMIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEAh4W6vcB9\nBqsdG6EQq6Lub/DGavHgVkmust1R9hmoaux/eQr9zMvgw8H+dwf0WDiwcD7xTUB4\niN8IPCUfhJL2jv413uPVsKSAwiEMWFerTMYV/hWWIY486daglcaUF8SLMtn6ool6\nF63IbcxGAxifqwKOuFyL5hZll3uupxqRJmdLpz7q9KZgLY0tIEiW3JqYuPHEZne/\nR6VrcPN0QR7MwS4mHX6H+s45UpJNLLA7PiINAyCFyBkfJfMUoa8KsISzHfMe+gHa\nhdayfrGCLgJRas1Mbo9gsqJAsd1SD6qURzl9IJuYO3TymurwuhFv/gFNkUmSgJAL\na40pFsveMG/xBwIDAQABo4ICvTCCArkwggFfBgNVHSAEggFWMIIBUjCCAU4GCCsG\nAQQBw1IFMIIBQDA9BggrBgEFBQcCARYxaHR0cHM6Ly93d3cuZS1jZXJ0Y2hpbGUu\nY2wvcG9saXRpY2FzLXktcHJhY3RpY2FzLzCB/gYIKwYBBQUHAgIwgfEege4ARQBs\nACAAcgBlAHMAcABvAG4AZABlAHIAIABlAHMAdABlACAAZgBvAHIAbQB1AGwAYQBy\nAGkAbwAgAGUAcwAgAHUAbgAgAHIAZQBxAHUAaQBzAGkAdABvACAAaQBuAGQAaQBz\nAHAAZQBuAHMAYQBiAGwAZQAgAHAAYQByAGEAIABkAGEAcgAgAGkAbgBpAGMAaQBv\nACAAYQBsACAAcAByAG8AYwBlAHMAbwAgAGQAZQAgAGMAZQByAHQAaQBmAGkAYwBh\nAGMAaQDzAG4ALgAgAFAAbwBzAHQAZQByAGkAbwByAG0AZQBuAHQAZQAsMB0GA1Ud\nDgQWBBQGqD9EUbhT12MOh8n+lbjhjogM9DALBgNVHQ8EBAMCBPAwIwYDVR0RBBww\nGqAYBggrBgEEAcEBAaAMFgowNDY3NTUxMC0zMB8GA1UdIwQYMBaAFHTWIbP0Wugt\nfLtZBmND72m0OpMEMEAGA1UdHwQ5MDcwNaAzoDGGL2h0dHA6Ly9jcmwuZWNlcnRj\naGlsZS5jbC9FLUNFUlRDSElMRUNBRkVTMDIuY3JsMD0GCCsGAQUFBwEBBDEwLzAt\nBggrBgEFBQcwAYYhaHR0cDovL29jc3BmZXMuZWNlcnRjaGlsZS5jbC9vY3NwMDwG\nCSsGAQQBgjcVBwQvMC0GJSsGAQQBgjcVCIHd0k2X1DqGiZ0Gh4jOb4OvzGBFhtLs\nCIbHyn0CAWQCAQcwIwYDVR0SBBwwGqAYBggrBgEEAcEBAqAMFgo5NjkyODE4MC01\nMA0GCSqGSIb3DQEBCwUAA4ICAQAsBd5o8R+gI2ExvfiMJWmXMEo8uohSExDSN6Z3\nPhkraKtDAr0WxuwWxMu13Zef5PogLqpL37vkrkwMFZmvZ6WVDLl2kg4nTULNyMDJ\njNMPB8aECvL1ddPFMYmbGna6ZjncAVN8pcau3RPmRdI/aDajx3xPTHuiLcITStK3\ndHUPIr3nC67YRzf/VCkGN3TpKcJ/wL1ct1XcL16yANSNGqc3PI9xdyCzpqH4Yj3y\nSGCYjd6JimA51+VyCMF7xEpFVxI3VrZS5UBcnHFra4Sv4LKQ0ae4WtkNZ9RsHrHN\nkHufWSRTLMn16729vKtEUFRlvYA5fMdSVI3II3Lrm52WybWm8MBOB0NwHDujWk3h\n9qnkNgqmLPzrY1/EcPCDW8e+VsQfCPPNxssUswFItWZyDjTRVr4+3eDNVQE28q/3\nhDWOiclyn1SuGvdnv0zBJak7bKUN8UUo84OEGXTD4/jNnqo9dt+/0ULT6FmBYjgj\n/nqFrHkfLjCcXJEUlLo45W4ovFfLO4j0iShHSNjk37B/1dhtnLMAfH0aB2zu3Qkf\nmXFRcRmh61gE2A/N9CuWSkgBGYgxRkYhRzRrDTxcs4tpb4D/xzP7Ok40PMXLE1Pu\nnhRkVaV8W9hvqhcEBgxA4xzY3t9t2BOVgJLDnqJzsoRmmGQx2dZYTB/5ccRG6PG4\n4iOVPw=='}}}}}, 'Signature': {'@xmlns': 'http://www.w3.org/2000/09/xmldsig#', 'SignedInfo': {'CanonicalizationMethod': {'@Algorithm': 'http://www.w3.org/TR/2001/REC-xml-c14n-20010315'}, 'SignatureMethod': {'@Algorithm': 'http://www.w3.org/2000/09/xmldsig#rsa-sha1'}, 'Reference': {'@URI': '#ID17424dbf30f0438eac7e555dded90038', 'Transforms': {'Transform': {'@Algorithm': 'http://www.w3.org/TR/2001/REC-xml-c14n-20010315'}}, 'DigestMethod': {'@Algorithm': 'http://www.w3.org/2000/09/xmldsig#sha1'}, 'DigestValue': 'avlhJI4n+Gvt6KTH5oeDv1id7iw='}}, 'SignatureValue': 'IoH7og/OkZ1SxuArA4/xsnLPkd3KtzjckazCjkjWp+WnveU52+b06e2Ko1mQ4xKB\nO1F9G6xFCYlgwYUXP0s8QnLy7J2E8143pAZK82K4W7KRzUA2RIfQegfI+9rzOb0W\n5LBj/0sOT3H43KNFnSePif7M3OGaJcdEb5A8S6sbS13iID/Xy3vcgMU4StAgf5S4\nXILzTB68tjD76v6sremmpXO+HIjkuJl0R5cycspM98wrysDvPaXDNk6WFw87atZ5\n4bp0JfdOQRGv4zUdPWdjBDczfW458+Fg13r49ckLbbQA1zNnWxZgnl0CfY1GmUIk\n8PIzTdnk54QEo/g6LUipCA==', 'KeyInfo': {'KeyValue': {'RSAKeyValue': {'Modulus': 'mZyNOZGHPazOujYza5bj8qjBUwShdlmdABF+CJO5S1iGAMBXoyb+UTc5tqKaqLQr\nhxOJxZSZwpuMVbc2cJ+JtDwXsEj2asqB9GXjHGUH4HhJu7PI7qWw2YivtQm2p4R3\nhsd+un1QECsfAyL8DQ4v8ulm7URFuOimYA8YK2ML8EmiBJ2AzXu9MTxIUBUkhxsv\nA3tffFuDudOtIK12FpHSM0jseAFvZJY7HSfV/oybdeUDeKX10rrgnIrsOHfwZQ9E\nTQ3lzbTZms1WCu40cG30aMn1o8c1TLMdfNPT8f2PULshH4TuykwHli4MVk8LD2TR\n4hDdNt6JK2BUQJT2VeL8Fw==', 'Exponent': 'AQAB'}}, 'X509Data': {'X509Certificate': 'MIIGEDCCBPigAwIBAgIINo2mVCGpjRYwDQYJKoZIhvcNAQELBQAwgboxHjAcBgkq\nhkiG9w0BCQEWD3NvcG9ydGVAaWRvay5jbDEiMCAGA1UEAwwZSURPSyBGSVJNQSBF\nTEVDVFJPTklDQSBWMjEXMBUGA1UECwwOUlVULTc2NjEwNzE4LTQxIDAeBgNVBAsM\nF0F1dG9yaWRhZCBDZXJ0aWZpY2Fkb3JhMRkwFwYDVQQKDBBCUE8gQWR2aXNvcnMg\nU3BBMREwDwYDVQQHDAhTYW50aWFnbzELMAkGA1UEBhMCQ0wwHhcNMjIwNTIzMjI0\nOTM1WhcNMjUwNTIyMjI0OTM1WjB9MSgwJgYDVQQDDB9DUklTVElBTiBVTElTRVMg\nUk9KQVMgR1VUSUVSUkVaMSIwIAYJKoZIhvcNAQkBFhNjcm9qYXNAZ2RleHByZXNz\nLmNsMRMwEQYDVQQFEwoxMDk3NDM3Ny0xMQswCQYDVQQGEwJDTDELMAkGA1UEBwwC\nUk0wggEiMA0GCSqGSIb3DQEBAQUAA4IBDwAwggEKAoIBAQCZnI05kYc9rM66NjNr\nluPyqMFTBKF2WZ0AEX4Ik7lLWIYAwFejJv5RNzm2opqotCuHE4nFlJnCm4xVtzZw\nn4m0PBewSPZqyoH0ZeMcZQfgeEm7s8jupbDZiK+1CbanhHeGx366fVAQKx8DIvwN\nDi/y6WbtREW46KZgDxgrYwvwSaIEnYDNe70xPEhQFSSHGy8De198W4O5060grXYW\nkdIzSOx4AW9kljsdJ9X+jJt15QN4pfXSuuCciuw4d/BlD0RNDeXNtNmazVYK7jRw\nbfRoyfWjxzVMsx1809Px/Y9QuyEfhO7KTAeWLgxWTwsPZNHiEN023okrYFRAlPZV\n4vwXAgMBAAGjggJUMIICUDAJBgNVHRMEAjAAMB8GA1UdIwQYMBaAFPBsM7+sl5NY\neqHgzp7s6N77ZT76MIGYBgNVHSAEgZAwgY0wgYoGCisGAQQBg4weAQQwfDAsBggr\nBgEFBQcCARYgaHR0cHM6Ly9wc2MuaWRvay5jbC9vcGVuL2Nwcy5wZGYwTAYIKwYB\nBQUHAgIwQB4+AEMAZQByAHQAaQBmAGkAYwBhAGQAbwAgAHAAYQByAGEAIAB1AHMA\nbwAgAFQAcgBpAGIAdQB0AGEAcgBpAG8wggEPBgNVHR8EggEGMIIBAjCB/6A6oDiG\nNmh0dHBzOi8vcHNjLmlkb2suY2wvb3Blbi9JRE9LX0ZJUk1BX0VMRUNUUk9OSUNB\nX1YyLmNybKKBwKSBvTCBujEeMBwGCSqGSIb3DQEJARYPc29wb3J0ZUBpZG9rLmNs\nMSIwIAYDVQQDDBlJRE9LIEZJUk1BIEVMRUNUUk9OSUNBIFYyMRcwFQYDVQQLDA5S\nVVQtNzY2MTA3MTgtNDEgMB4GA1UECwwXQXV0b3JpZGFkIENlcnRpZmljYWRvcmEx\nGTAXBgNVBAoMEEJQTyBBZHZpc29ycyBTcEExETAPBgNVBAcMCFNhbnRpYWdvMQsw\nCQYDVQQGEwJDTDAdBgNVHQ4EFgQUnR78AetPsNDqFh+h8AeExJY4aOgwCwYDVR0P\nBAQDAgSQMCMGA1UdEgQcMBqgGAYIKwYBBAHBAQKgDBYKNzY2MTA3MTgtNDAjBgNV\nHREEHDAaoBgGCCsGAQQBwQEBoAwWCjEwOTc0Mzc3LTEwDQYJKoZIhvcNAQELBQAD\nggEBAA6KTv23rQSdvQrJMy1jxE/+gYgMDsPqx6VcSRrsDVl+tUjf4Bld1zBLmBak\ndtMPiyNhQ0kaOgEjo3QU8kQ/SV6fWysnmwwAutagLJvX5cix9YPrhAnGxe31kdR7\nnj8h/xMTetxxgmOQ/+sKwM6GDPCyzVMZ0JuXr9rn3ozViDx0+Lu1tegCE0CMZgLi\nynwZXrtR5bjbJH01QrxErY8GoFIY7BO8Iah/iBS0SfClWYaEH6JOjcGUSIwDapa3\nTh0+GYEUgSektb8aHqyl2XEDJtAem4PSWmsdOBZZaXA07eUVxI20qq4FeKdl38Mi\nt/WiFQvQ0chqv6iEPJZqx/3FVIU='}}}}}
    
    """
    #print(xml_data)
    #pause=input("Presione una tecla para continuar...")
    
    

    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '33':
            tipoDocumento = "FEAV"
            envioTipo = "EnvioDTE"
            accountNumber = "1110401021"
            price = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Totales']['MntNeto']
    except:
        pass
    try:
        if xml_data['EnvioBOLETA']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '39':
            tipoDocumento = "BEAV"
            envioTipo = "EnvioBOLETA"
            accountNumber = "1110401023"
            price = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Totales']['MntTotal'] #cambiar por el total
    except:
        pass 
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '34':
            tipoDocumento = "FEEV"
            envioTipo = "EnvioDTE"
    except:
        pass
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '61':
            tipoDocumento = "NCAV"
            envioTipo = "EnvioDTE"
    except:
        pass
    try:
        if xml_data['EnvioDTE']['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['TipoDTE'] == '56':
            tipoDocumento = "NDAV"
            envioTipo = "EnvioDTE"
    except:
        pass    

    


    #print(envioTipo)

    fechaEmis = datetime.strptime(xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['FchEmis'], "%Y-%m-%d")
    print(fechaEmis)
    
    fechaVenc = fechaEmis + timedelta(days=30)
    print(fechaVenc)
    try:
        
        contactIndex = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['DirRecep']
        
    except:
        contactIndex = 'noAddress'


    try:
        giro = xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Emisor']['GiroRecep']
    except:
        giro = 'noGiroReceptor'
    
    try:
        nroConfirmacion = next(c['#text'] for c in xml_data['EnvioDTE']['Personalizados']['DocPersonalizado']['campoNumero'] if c['@name']=='NroReserva')

    except:
        nroConfirmacion = '0'
    
    



    payload = {
    "documentType": tipoDocumento, # OBLIGATORIO, CORRESPONDE AL TIPO DE DOCUMENTO
    "firstFolio": xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['Folio'], # OBLIGATORIO, CORRESPONDE AL NÚMERO DE DOCUMENTO (SI SE DEJA EN 0, TOMARÁ EL CORRELATIVO)
    #"lastFolio": xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['IdDoc']['Folio'], # OBLIGATORIO, CORRESPONDE AL NÚMERO DE DOCUMENTO (SI SE DEJA EN 0, TOMARÁ EL CORRELATIVO)
    "externalDocumentID": "", # SE PUEDE ENVIAR VACÍO, ID EXTERNA PARA IDENTIFICAR EL DOCUMENTO EN CASO DE QUE SE NECESITE
    "emissionDate": { # OBLIGATORIA, FECHA DE EMISIÓN
        "day": fechaEmis.day,
        "month": fechaEmis.month,
        "year": fechaEmis.year
    },
    "firstFeePaid": { # OBLIGATORIA, FECHA DEL PRIMER PAG
        "day": fechaVenc.day,
        "month": fechaVenc.month,
        "year": fechaVenc.year
    },
    "clientFile": formatear_rut(xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Receptor']['RUTRecep']),     #  OBLIGATORIA, ID DE FICHA DEL CLIENTE
    "contactIndex": contactIndex,   #  OBLIGATORIA, DIRECCIÓN CLIENTE
    "paymentCondition": "CREDITO30",  # OBLIGATORIA, CONDICIÓN DE PAGO
    "sellerFileId": "MarinaDunas",   # ID DEL VENDEDOR, OBLIGATORIO
    "clientAnalysis": {          # ANALISIS DE CUENTA CONTABLE POR ASIENTO DEL CLIENTE
        "accountNumber": accountNumber,  # NÚMERO DE CUENTA CONTABLE DEL ASIENTO POR CLIENTES
        "businessCenter": "EMPDNSADM000000", # CENTRO DE NEGOCIOS EN CASO DE QUE LA CUENTA ESTÉ CONFIGURADA PARA USAR CENTRO DE NEGOCIOS.
        "classifier01": "", # CLASIFICADOR1 EN CASO DE QUE LA CUENTA ESTÉ CONFIGURADA PARA USAR CLASIFICADORES.
        "classifier02": "" # CLASIFICADOR2 EN CASO DE QUE LA CUENTA ESTÉ CONFIGURADA PARA USAR CLASIFICADORES.
    },
    "billingCoin": "Peso",    # OBLIGATORIO, ID DE MONEDA A UTILIZAR
    "billingRate": 1,     # OBLIGATORIDE CAMBIO DE MONEDA
    "shopId": "Marina Dunas",   # ID DE LOCAL, OBLIGATORIO
    "priceList": "1", # LISTA DE PRECIOS, OBLIGATORIA
    "giro": giro,    # OBLIGATORIO, GIRO DEL DOCUMENTO
    "district": xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Emisor']['CiudadOrigen'],  # COMUNA, OBLIGATORIA
    "city": xml_data[envioTipo]['SetDTE']['DTE']['Documento']['Encabezado']['Emisor']['CiudadOrigen'],   # OBLIGATORIA, CIUDAD
    "contact": -1,  # ENVIAR COMO -1
    "attachedDocuments": [   # SE PUEDE DEJAR VACÍO EN CASO DE QUE NO EXISTAN DOCUMENTOS ASOCIADOS (ENVIAR COMO []))
        #{
        #"date": {
        #    "day": 0,
        #    "month": 0,
        #    "year": 0
        #},
        #"documentTypeId": "",	# tipo de documento codigo
        #"folio": "",	# folio
        #"reason": ""	# comentario de documento
        #}
    ],
    "storage": {
        "code": "",   # CODIGO DE BODEGA (OBLIGATORIO EN CASO DE QUE SE MUEVA STOCK)
        "motive": "", # MOTIVO DE MOVIMIENTO DE STOCK (OBLIGATORIO EN CASO DE QUE SE MUEVA STOCK)
        "storageAnalysis": {  # ANALISIS POR CUENTA CONTABLE DE INVENTARIO
        "accountNumber": "",
        "businessCenter": "",
        "classifier01": "",
        "classifier02": ""
        }
    },
    "details": [   # OBLIGATORIO, DETALLES DE LA GUÍA DE DESPACHO.
        {
        "type": "S", # TIPO DE PRODUCTO (PUEDE SER "A" SI ES UN ARTICULO O "S" SI ES UN SERVICIO) (OBLIGATORIA)
        "isExempt": False, # SE DEFINE SI ES EXENTA O NO (OBLIGATORIO)
        "code": "Alojamiento", # CODIGO DEL PRODUCTO
        "count": 1,  # CANTIDAD
        "productName": "Alojamiento", # NOMBRE DEL PRODUCTO
        "productNameBarCode": nroConfirmacion, # CODIGO DE BARRAS DEL PRODUCTO
        "price": price, # PRECIO
        "discount": { # DESCUENTOS POR LINEA DE DETALLE EN CASO DE QUE EXISTAN (EN CASO DE QUE NO EXISTAN, ENVIAR AMBOS DATOS COMO 0)
            "type": 0,
            "value": 0
        },
        "unit": "Unidad",  # UNIDAD DEL PRODUCTO
        "analysis": {  # ANALISIS DE CUENTA CONTABLE POR ASIENTO DE VENTAS
            "accountNumber": "3110101002", #3110101002
            "businessCenter": "EMPDNSADM000000",
            "classifier01": "",
            "classifier02": ""
        },
        "useBatch": False,  # EN CASO DE USAR LOTES, SE ENVÍA COMO TRUE, SINO FALSE
        "batchInfo": [  # EN CASO DE NO USAR LOTES, ENVIAR COMO UN ARREGLO VACÍO []
            #{
            #"amount": 0,
            #"batchNumber": ""
            #}
        ]
        }
    ],
    "saleTaxes": [  # OBLIGATORIO, ANALISIS DE CUENTA CONTABLE POR ASIENTO DE IMPUESTOS
        {
        "code": "IVA", # CODIGO DEL IMPUESTO
        "value": 19, # VALOR DEL IMPUESTO
        "taxeAnalysis": {
            "accountNumber": "2120301001",
            "businessCenter": "EMPDNSADM000000",
            "classifier01": "",
            "classifier02": ""
        }
        }
    ],
    "ventaRecDesGlobal": [  # OPCIONAL, DESCUENTOS GLOBALES, EN CASO DE NO USAR ENVIAR COMO UN ARREGLO VACÍO []
        #{
        #"amount": 0,
        #"modifierClass": "",
        #"name": "",
        #"percentage": 0,
        #"value": 0
        #}
    ],
    "gloss": f"Comentario: \n Opera Confirmation No: {nroConfirmacion}",  # GLOSA DEL DOCUMENTO
    #"customFields": [ # OPCIONAL, CAMPOS PERSONALIZABLES, EN CASO DE NO USAR ENVIAR COMO UN ARREGLO VACÍO []
    #    {
    #    "name": "NroReserva",
    #    "value": nroConfirmacion
    #    }
    #],
    "isTransferDocument": True  # DEFINE SI SERÁ UN DOCUMENTO DE TRASPASO O NO (SI ES FALSE SE ENVÍA AL SII, SI ES TRUE NO SE ENVÍA AL SII).
    }                              
    print(payload)
    return payload

def guardar_venta(payload):
    """
    Envía una solicitud para guardar una venta en Defontana usando un diccionario de datos.
    Args:
        venta (dict): Diccionario con los campos requeridos por la API.

    Returns:
        dict o str: Respuesta de la API.

    
    """
    url = f"https://{apiValue}.defontana.com/api/Sale/SaveSale"
    headers = get_auth_headers()
    if not headers:
        print("No se pudo obtener un token válido para autenticar la petición.")
        return None
    try:
        print(f"\n Intentando subir el siguiente payload: \n{payload}")
        resp = requests.post(url, headers={**headers, "Content-Type": "application/json"}, json=payload)
        #print("\n Respuesta del servidor (guardar_venta):")
        return resp
            
    except Exception as e:
        print(f"\n Excepción en guardar_venta(): {e}")
        return None
        







#funciones de la API de Defontana para Vouchers Contables
def importar_arqueo_opera(filename):
    """
    Procesa un archivo Excel de arqueo, lee sus datos y los inserta en una base de datos
    si no existen previamente.
    
    Args:
        filename (str): Ruta al archivo Excel a procesar
        
    Returns:
        list: Mensajes informativos del proceso
        dict: Estadísticas del procesamiento (registros procesados, insertados, duplicados, errores)
    """
    mensajes = []
    estadisticas = {
        "procesados": 0,
        "insertados": 0,
        "duplicados": 0,
        "errores": 0
    }
    errores = []
    try:
        # Registrar inicio del proceso
        mensajes.append(f"Iniciando procesamiento del archivo: {filename}")
        print(f"Iniciando procesamiento del archivo: {filename}")
        
        # Abrir archivo Excel
        try:
            workbook = openpyxl.load_workbook(filename, read_only=True)  # Modo read_only para archivos grandes
        except Exception as e:
            mensaje_error = f"Error al abrir el archivo Excel: {e}"
            mensajes.append(mensaje_error)
            print(colored(f" {ERROR_SYMBOL} {mensaje_error}", "red"))
            return mensajes, None
            
        # Conectar a la base de datos
        try:
            conexion = connect()
        except Exception as e:
            mensaje_error = f"Error de conexión a la base de datos: {e}"
            mensajes.append(mensaje_error)
            print(colored(f" {ERROR_SYMBOL} {mensaje_error}", "red"))
            workbook.close()
            return mensajes, None
            
        # Obtener hoja activa
        hoja = workbook.active
        mensajes.append(f"Archivo abierto correctamente. Procesando datos...")
        print(colored(f"Archivo abierto correctamente. Procesando datos...", "white"))
        
        # Procesar filas del archivo
        for i, fila in enumerate(hoja.iter_rows(min_row=3), start=3):
            try:
                # Obtener valores de la fila
                valores = [celda.value for celda in fila]
                
                # Verificar si la fila está vacía o es un encabezado
                if len(valores) < 15 or valores[1] == 'TRX_DATE' or valores[1] is None:
                    continue
                
                
                estadisticas["procesados"] += 1
                
                # Crear el índice único
                index_column = f"{valores[0]}_{valores[1]}_{valores[2]}_{valores[8]}_{valores[14]}"
                
                # Normalizar valores
                trx_code = valores[0]
                trx_date = datetime.strptime(valores[1], "%d/%m/%Y %H:%M:%S")
                
                
                pname = valores[3] or ""
                room = valores[6]
                remark = valores[8] or ""
                reference = valores[11] or ""
                amt = valores[12]
                confirmation_no = valores[14]
                bill_no = valores[16]
                folio_type = valores[19]
                fiscal_bill_no = valores[21]
                cashier_id = valores[23]
                app_user = valores[24] or ""
                
                # Verificar si el registro ya existe
                print(colored(f"{ASK_SYMBOL} Consultando si existe el registro en la base de datos...", "white"))
                query = f"SELECT * FROM arqueo WHERE index_column = '{index_column}'"
                resultado = ejecutar_consulta(conexion, query)
                
                if resultado:
                    estadisticas["duplicados"] += 1
                    print(colored(f" {WARNING_SYMBOL} El registro ya existe en la base de datos, no se insertará nuevamente.", "yellow"))
                    continue
                
                # Preparar query de inserción más segura
                query = """
                    INSERT INTO arqueo (
                        trx_code, trx_date, pname, room, remark, reference, amt,
                        confirmation_no, bill_no, folio_type, fiscal_bill_no,
                        cashier_id, app_user, index_column
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                # Insertar el registro usando parámetros en lugar de concatenación
                valores_insercion = (
                    trx_code, trx_date, pname, room, remark, reference, amt,
                    confirmation_no, bill_no, folio_type, fiscal_bill_no,
                    cashier_id, app_user, index_column
                )
                #print(valores_insercion)
                insertar_resultados_seguros(conexion, query, valores_insercion)
                print(colored(f" {OK_SYMBOL} Registro insertado correctamente.", "green"))
                estadisticas["insertados"] += 1
                
            except Exception as e:
                estadisticas["errores"] += 1
                mensaje_error = f"Error procesando fila {i}: {e}"
                mensajes.append(mensaje_error)
                print(colored(f" {ERROR_SYMBOL} {mensaje_error}", "red"))
                errores.append(mensaje_error)
                
        # Resumen de procesamiento
        mensaje_resumen = (f"Procesamiento completado: {estadisticas['procesados']} registros procesados, "
                            f"{estadisticas['insertados']} insertados, {estadisticas['duplicados']} duplicados, "
                            f"{estadisticas['errores']} errores.")
        mensajes.append(mensaje_resumen)
        print(colored(mensaje_resumen, "green"))
        
        # Cerrar recursos
        workbook.close()
        conexion.close()
        
        return mensajes, estadisticas
        
    except FileNotFoundError:
        mensaje_error = f"Error: Archivo no encontrado: {filename}"
        mensajes.append(mensaje_error)
        print(colored(f" {ERROR_SYMBOL} {mensaje_error}", "red"))
        errores.append(mensaje_error)
        return mensajes, None
        
    except Exception as e:
        mensaje_error = f"Error general: {e}"
        mensajes.append(mensaje_error)
        print(colored(f" {ERROR_SYMBOL} {mensaje_error}", "red"))
        errores.append(mensaje_error)
        return mensajes, None
    
    finally:
        if len(errores) > 0:
            print(colored(f" {ERROR_SYMBOL}, Errores:", "red"))
            for error in errores:
                print(colored(f" {ERROR_SYMBOL} {error}", "red"))
        workbook.close()
        conexion.close()


# Esta función es necesaria para reemplazar el insertar_resultados actual
# con una versión que use consultas parametrizadas
def insertar_resultados_seguros(conexion, query, valores):
    """
    Inserta registros en la base de datos de manera segura usando consultas parametrizadas.
    
    Args:
        conexion: Conexión a la base de datos
        query: Consulta SQL con placeholders (%s)
        valores: Tupla con los valores a insertar
    """
    cursor = conexion.cursor()
    cursor.execute(query, valores)
    conexion.commit()
    cursor.close()
    return True

def consultar_voucher(type, number, fiscal_year):
    payload = {
        "VoucherType": "type",
        "Number": number ,
        "FiscalYear": fiscal_year,
        "Isopening": False
    }
    

    url = f"https://api.defontana.com/api/Accounting/GetVoucher"
    headers = get_auth_headers()
    if not headers:
        print("No se pudo obtener un token válido para autenticar la petición.")
        return None
    
    try:
        resp = requests.get(url, headers=headers, params=payload)
        if resp.status_code == 200:
            #print("Respuesta GetClients:")
            data = resp.json()
            #print(data)
            return data
        else:
            print(f"Error al consultar voucher: {resp.status_code} {resp.text}")
            return None
    except Exception as e:
        print(f"Excepción en consultar_voucher(): {e}")
        return None
        print(f"Excepción en consultar_voucher(): {e}")
        return None

def crear_payload_voucher_lines(fecha_str_input):

    #fecha_str = fecha_str_input

    fecha_obj = datetime.strptime(fecha_str_input, '%Y-%m-%d')
    fecha_para_sql = fecha_obj.strftime('%Y-%m-%d')
    fecha_iso_8601 = fecha_obj.strftime('%Y-%m-%dT00:00:00.000Z')
    #print(f"Fecha formateada para SQL: {fecha_para_sql}")

    query_voucher = f"SELECT * FROM vouchercaja where fecha = '{fecha_para_sql}' order by cuenta asc;"    

    try: 
        consulta = ejecutar_consulta(connect(), query_voucher)
        
    except Exception as e:
        print(f"Error al ejecutar la consulta: {e}")
        return None
    
    #print(consulta)
    payload_header = {
            "fiscalYear": fecha_obj.year,
            "number": 0,                    #dejar en cero para que tome correlativo
            "voucherType": 'TRASPASO',
            "date": fecha_iso_8601,
            "comment": f"Caja del {fecha_obj.day} de {fecha_obj.month} del {fecha_obj.year}",
        }
    #print(f"\n Encabezado del Voucher:{payload_header}")
    
    payload_detalle = []
    no_linea_detalle = 0
    for fila in consulta:
        #print(fila)
        no_linea_detalle += 1
        
        doc_numero = 0
        if fila[9] and str(fila[9]).isdigit():
            doc_numero = int(fila[9])
        
        fecha_vencimiento = None
        if fila[10] and str(fila[10]).strip():  # Si hay fecha de vencimiento
            try:
                # Convertir a formato ISO 8601
                fecha_venc_obj = datetime.strptime(str(fila[10]), '%Y-%m-%d')
                fecha_vencimiento = fecha_venc_obj.strftime('%Y-%m-%dT00:00:00.000Z')
            except:
                fecha_vencimiento = None  # Si hay error, dejar en None
        
        linea_detalle = {
            "accountCode": str(fila[1]),
            "debit": int(fila[4]),
            "credit": int(fila[5]),
            "secondaryDebit": 0,
            "secondaryCredit": 0,
            "exchangeRate": 0,
            "comment": str(fila[3]),
            "fileId": fila[6],
            "documentType": str(fila[8]),
            "documentSeries": "",
            "documentNumber": doc_numero,
            "documentExpirationDate": fecha_vencimiento,
            "bussinessCenterId": "",
            "classifier1Id": "",
            "classifier2Id": "",
            "referenceCurrencyId": "",
            "referenceExchangeRate": 0,
            "movementTypeId": "",
            "movementSeries": "",
            "movementNumber": "",
            "accountAmountRate": 0,
            "ctaCreditOrDebitAmount": 0
            }
        #print(f"\n Linea detalle No: {no_linea_detalle}\n ->{linea_detaile}")
        payload_detalle.append(linea_detalle)

    #construir payload final
    payload_voucher= {
        "header": payload_header,
        "detail": payload_detalle,
        "automaticFoliation": True
    }
    return payload_voucher

def calcular_totales_voucher(payload_voucher):
    """
    Calcula los totales debe y haber de un voucher.
    
    Args:
        payload_voucher (dict): Diccionario con los datos del voucher
        
    Returns:
        tuple: (total_debe, total_haber)
    """
    total_debe = sum(item.get('debit', 0) for item in payload_voucher.get('detail', []))
    total_haber = sum(item.get('credit', 0) for item in payload_voucher.get('detail', []))
    return total_debe, total_haber

def verificar_voucher_existente(fecha, glosa):
    """
    Verifica si ya existe un voucher para una fecha y glosa específicas.
    
    Args:
        fecha (str): Fecha del voucher en formato 'YYYY-MM-DD'
        glosa (str): Glosa o comentario del voucher
        
    Returns:
        dict: Información del voucher si existe, None si no existe
    """
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        cursor = conn.cursor(dictionary=True)
        query = """
        SELECT * FROM vouchers_enviados 
        WHERE fecha = %s AND glosa = %s
        """
        cursor.execute(query, (fecha, glosa))
        result = cursor.fetchone()
        return result
    except Error as e:
        print(f"Error al verificar voucher existente: {e}")
        return None
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            
def registrar_voucher_enviado(payload_voucher, response_data):
    """
    Registra en la base de datos un voucher enviado exitosamente.
    
    Args:
        payload_voucher (dict): Datos del voucher enviado
        response_data (dict): Respuesta de la API con datos del voucher registrado
        
    Returns:
        bool: True si se registró correctamente, False en caso contrario
    """
    conn = get_db_connection()
    if not conn:
        return False
    
    try:
        cursor = conn.cursor()
        fecha = payload_voucher.get('header', {}).get('date')
        glosa = payload_voucher.get('header', {}).get('comment', '')
        total_debe, total_haber = calcular_totales_voucher(payload_voucher)
        
        query = """
        INSERT INTO vouchers_enviados 
        (fecha, glosa, total_debe, total_haber, numero_voucher, tipo_voucher, anio_fiscal, json_payload)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        cursor.execute(query, (
            fecha,
            glosa,
            total_debe,
            total_haber,
            response_data.get('number'),
            response_data.get('voucherType'),
            response_data.get('fiscalYear'),
            json.dumps(payload_voucher)
        ))
        
        conn.commit()
        return True
    except Error as e:
        print(f"Error al registrar voucher enviado: {e}")
        return False
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            
def mostrar_resultado(resultado):
    """
    Muestra el resultado de una operación de envío de voucher de forma amigable.
    
    Args:
        resultado (dict): Diccionario con el resultado de la operación
    """
    if resultado.get("already_exists"):
        print(f"El comprobante ya existe en la base de datos:")
        print(f"  - Número: {resultado.get('number')}")
        print(f"  - Tipo: {resultado.get('voucher_type')}")
        print(f"  - Año fiscal: {resultado.get('fiscal_year')}")
        print(f"  - Total Debe: ${resultado.get('total_debe'):,.2f}")
        print(f"  - Total Haber: ${resultado.get('total_haber'):,.2f}")
    elif resultado.get("success"):
        print(f"Comprobante enviado y registrado exitosamente:")
        print(f"  - Número asignado: {resultado.get('number')}")
        print(f"  - Tipo: {resultado.get('voucher_type')}")
        print(f"  - Año fiscal: {resultado.get('fiscal_year')}")
        print(f"  - Total Debe: ${resultado.get('total_debe'):,.2f}")
        print(f"  - Total Haber: ${resultado.get('total_haber'):,.2f}")
    else:
        print(f"Error al procesar el comprobante: {resultado.get('message')}")
            
            


def subir_voucher_defontana(payload_voucher):
    """
    Envía una solicitud para guardar un voucher en Defontana usando un diccionario de datos.
    Args:
        payload_voucher (dict): Diccionario con los campos requeridos por la API.

    Returns:
        dict: Diccionario con los detalles completos de la respuesta, incluyendo:
            - status_code: Código de estado HTTP
            - success: Indicador de éxito de la operación según la API
            - message: Mensaje de la API
            - exception_message: Mensaje de excepción si existe
            - voucher_type: Tipo de voucher
            - number: Número de voucher
            - fiscal_year: Año fiscal
            - raw_response: Respuesta completa sin procesar
            - error: Información de error en caso de excepción
    """
    url = "https://api.defontana.com/api/Accounting/InsertVoucher"
    headers = get_auth_headers()
    if not headers:
        print("No se pudo obtener un token válido para autenticar la petición.")
        return {
            "status_code": None,
            "success": False,
            "message": "No se pudo obtener un token válido para autenticar la petición.",
            "error": "Autenticación fallida"
        }
    
    try:
        print(f"\nEnviando solicitud a {url}")
        resp = requests.post(
            url, 
            headers={**headers, "Content-Type": "application/json"}, 
            json=payload_voucher  # Enviar directamente el payload sin envolver
        )
        
        print(f"Código de respuesta: {resp.status_code}")
        
        # Preparar el resultado base con el código de estado
        result = {
            "status_code": resp.status_code,
            "raw_response": resp.text
        }
        
        # Intentar extraer el JSON de la respuesta
        try:
            json_response = resp.json()
            # Añadir todos los campos del JSON al resultado
            result.update({
                "success": json_response.get("success"),
                "message": json_response.get("message"),
                "exception_message": json_response.get("exceptionMessage"),
                "voucher_type": json_response.get("voucherType"),
                "number": json_response.get("number"),
                "fiscal_year": json_response.get("fiscalYear")
            })
        except ValueError:
            # Si no se puede convertir a JSON, se deja solo el texto crudo
            result["success"] = False
            result["message"] = "No se pudo procesar la respuesta como JSON"
        
        if resp.status_code != 200:
            print(f"Error en la respuesta: {resp.text}")
            result["success"] = False
        else:
            print("Voucher subido exitosamente")
            if "success" not in result:
                result["success"] = True
        
        return result
            
    except Exception as e:
        print(f"\nExcepción en subir_voucher(): {e}")
        return {
            "status_code": None,
            "success": False,
            "message": f"Excepción: {str(e)}",
            "error": str(e)
        }
    


def procesar_voucher_dia(fecha):
    """
    Procesa y envía el voucher para una fecha específica
    
    Args:
        fecha (str): Fecha en formato 'YYYY-MM-DD'
    """
    # Generar el payload
    payload = crear_payload_voucher_lines(fecha)
    
    if not payload:
        print(f"No se pudo crear el payload para la fecha {fecha}")
        return
    
    # Validar que el payload esté balanceado (débitos = créditos)
    total_debito = sum(item["debit"] for item in payload["detail"])
    total_credito = sum(item["credit"] for item in payload["detail"])
    
    fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
    fecha_formato = fecha_obj.strftime('%d/%m/%Y')
    
    # Visualización en formato de comprobante contable
    print("\n" + "="*80)
    print(f"{'COMPROBANTE DE TRASPASO':^80}")
    print(f"{'FECHA: ' + fecha_formato:^80}")
    print(f"{'COMENTARIO: ' + payload['header']['comment']:^80}")
    print("="*80)
    print(f"{'CUENTA':<12} {'DESCRIPCIÓN':<40} {'DEBE':>12} {'HABER':>12}")
    print("-"*80)
    
    # Detalles
    for item in payload["detail"]:
        cuenta = item["accountCode"]
        comentario = item["comment"][:38] + '..' if len(item["comment"]) > 40 else item["comment"]
        debe = f"${item['debit']:,.0f}" if item['debit'] > 0 else ""
        haber = f"${item['credit']:,.0f}" if item['credit'] > 0 else ""
        
        print(f"{cuenta:<12} {comentario:<40} {debe:>12} {haber:>12}")
    
    print("-"*80)
    print(f"{'TOTALES':^52} ${total_debito:,.0f} ${total_credito:,.0f}")
    print("="*80)
    
    # Validación de balance
    if total_debito != total_credito:
        print(f"\n⚠️  ¡ADVERTENCIA! El voucher NO está balanceado:")
        print(f"    Débitos: ${total_debito:,.0f}")
        print(f"    Créditos: ${total_credito:,.0f}")
        print(f"    Diferencia: ${total_debito - total_credito:,.0f}")
    else:
        print(f"\n✅ Voucher balanceado correctamente: Total=${total_debito:,.0f}")
    
    # Opciones para el usuario
    print("\nOpciones:")
    print("1. Enviar comprobante a Defontana")
    print("2. Ver JSON completo")
    print("3. Cancelar")
    
    opcion = input("\nSeleccione una opción (1-3): ")
    
    if opcion == "1":
        # Enviar a Defontana
        print("\nEnviando comprobante a Defontana...")
        respuesta = subir_voucher_defontana(payload)
        if respuesta["success"]:
            print(f"Voucher creado exitosamente con número: {respuesta['number']}")
        else:
            print(f"Error: {respuesta['message']}")
        return respuesta
    elif opcion == "2":
        # Mostrar JSON completo
        import json
        print("\nJSON completo del comprobante:")
        print(json.dumps(payload, indent=2, ensure_ascii=False))
        
        # Preguntar si quiere enviar después de ver el JSON
        if input("\n¿Enviar este comprobante a Defontana ahora? (s/n): ").lower() == 's':
            print("\nEnviando comprobante a Defontana...")
            respuesta = subir_voucher_defontana(payload)
            return respuesta
        else:
            print("Envío cancelado por el usuario.")
            return None
    else:
        print("Operación cancelada por el usuario.")
        return None






#funciones para ejecutar desde la terminal

def procesar_fiscal_opera(filepath):
    """
    Procesa el archivo fiscal y genera (yields) mensajes de progreso.
    """
    conexion = None # Inicializa la conexión fuera del try para el finally
    try:
        # Intenta conectar a la base de datos primero
        print( "Conectando a la base de datos...")
        conexion = connect()
        if not conexion or not conexion.is_connected():
            print( "Error: No se pudo conectar a la base de datos.")
            return # Termina la ejecución si no hay conexión

        print( "Conexión a la base de datos establecida.")
        print( "Abriendo el archivo Excel...")
        workbook = openpyxl.load_workbook(filepath)
        #print( f"Archivo {filepath.filename} abierto. Procesando hojas...") 
        hojas = workbook.sheetnames
        hoja = workbook.active
        total_rows = hoja.max_row - 2 # Asumiendo que los datos empiezan en la fila 3
        processed_rows = 0

        print( f"Procesando {total_rows} filas...")

        for i, fila in enumerate(hoja.iter_rows(min_row=3), start=1):  # Comenzar desde la fila 3
            valores = [celda.value for celda in fila]

            # Validar si la fila está vacía o incompleta (ejemplo básico)
            if not any(valores) or valores[1] is None or valores[2] is None or valores[6] is None:
                print( f"Advertencia: Saltando fila {i+2} por datos faltantes o vacía.")
                continue

            fecha = valores[0]
            tipo = int(valores[1]) # Asegurar que sea int para la API
            folio = int(valores[2]) if valores[2] is not None else 0 # Asegurar que sea int
            anticipo = valores[3] if valores[3] is not None else ''
            conf_no = valores[4]
            razon_social = valores[5] if valores[5] else 'S/Razón Social' # Evitar None
            rut_raw = valores[6]
            rut = formatRut(str(rut_raw)) if rut_raw is not None else '' # Convertir a str antes de formatear
            neto = valores[7] if valores[7] is not None else 0
            exento = valores[8] if valores[8] is not None else 0
            iva = valores[9] if valores[9] is not None else 0
            total = valores[10] if valores[10] is not None else 0
            cajero = valores[11]
            index = f"{tipo}_{folio}"
            origen = 'Opera'

            # Mensaje de progreso por fila
            print( colored(f" \n{NEW_SYMBOL} Procesando fila {i}/{total_rows}: DTE {tipo}-{folio}, RUT {rut}, Total {total}","blue"))

            try:
                # Consultar estado SII (puede tardar)
                print( f"  -> Consultando estado SII para {tipo}-{folio}...")
                # Usar el rutEmisor definido en gde.py o pasarlo como argumento si varía
                estadoSii_response = consultaEstado(tipo, folio, rutEmisor)
                estadoSii = estadoSii_response.get('Data', 'Error en respuesta API') # Obtener Data o mensaje de error
                print( f"  -> Estado SII: {estadoSii}")

                # Consultar PDF (puede tardar)
                print( f"  -> Obteniendo PDF para {tipo}-{folio} en GDE...")
                pdf_response = consultaPDF(tipo, folio, rutEmisor)
                pdf_data = pdf_response.get('Data', 'Error en respuesta API') # Obtener Data o mensaje de error
                # Considera no incluir el PDF completo en el mensaje de progreso si es muy grande
                print( f"  -> PDF obtenido correctamente...)" )

                # Consultar XML (puede tardar)
                payload_xml = { 
                    "Environment" : "P",            # Se usará T para la Homologación y P para Producción.
                    "Group" : "E",                  # Se usará E para Emitidos y R para Recibidos
                    "Rut" : "76708884-1", 
                    "DocType" : tipo,               # 33=Factura, 39=Boleta, 41=Nota de Crédito, 52=Nota de Débito, 61= Nota de Crédito Electrónica, 56=Factura Electrónica, 61=Factura Exenta Electrónica, 52=Factura Exenta Electrónica, 110=Factura Exportación Electrónica
                    "Folio" : folio, 
                    "IsForDistribution" : "false" 
                    }
                
                print( f"  -> Obteniendo XML para {tipo}-{folio} en GDE...")
                try:
                    xml_data = recuperar_xml(payload_xml)['RecoverDocumentResult']['Data']
                    print( f"  -> XML codificado: {xml_data[:50]}..." ) # Solo muestra los primeros 50 caracteres
                except Exception as e:
                    print( f"  -> Error decodificando XML: {e}")
                    xml_decodificado = None
                

                # Inserción en la base de datos
                #print(fecha)
                query = f"""
                    INSERT IGNORE INTO fiscales
                    (`fecha`, `tipo`, `folio`, `anticipo`, `conf_no`, `razon_social`, `rut`, `neto`, `exento`, `iva`, `total`, `cajero`, `index`, `estadoSii`, `pdf`, `xml`,`origen`)
                    VALUES
                    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)
                """
                # Usar parámetros previene inyección SQL y maneja tipos de datos
                params = (
                    fecha, tipo, folio, anticipo, conf_no, razon_social, rut,
                    neto, exento, iva, total, cajero, index, estadoSii, pdf_data, xml_data, origen
                )

                print( f"  -> Insertando DTE {tipo}-{folio} en la base de datos...")
                cursor = conexion.cursor()
                cursor.execute(query, params)
                conexion.commit()
                rows_affected = cursor.rowcount
                cursor.close()
                if rows_affected > 0:
                    print( f"  -> DTE {tipo}-{folio} insertado correctamente.")
                else:
                    print( f"  -> DTE {tipo}-{folio} ya existía (o error al insertar).")
                processed_rows += 1
                
            except requests.exceptions.RequestException as api_err:
                print( f"  -> Error de API consultando {tipo}-{folio}: {api_err}")
            except mysql.connector.Error as db_err:
                print( f"  -> Error de Base de Datos insertando {tipo}-{folio}: {db_err}")
                conexion.rollback() # Revertir si falla la inserción
            except Exception as e:
                print( f"  -> Error inesperado procesando fila {i+2} (DTE {tipo}-{folio}): {e}")

            # Pequeña pausa opcional para que el navegador pueda renderizar las actualizaciones
            # time.sleep(0.05)

        workbook.close()
        print( f"Procesamiento completado. {processed_rows} de {total_rows} filas procesadas e intentadas insertar.")

    except FileNotFoundError:
        print( f"Error: Archivo no encontrado.")
    except openpyxl.utils.exceptions.InvalidFileException:
        print( "Error: El archivo subido no parece ser un archivo Excel válido (.xlsx).")
    except mysql.connector.Error as err:
        print( f"Error de conexión inicial a la base de datos: {err}")
    except Exception as e:
        print( f"Error general durante el procesamiento: {e}")
    finally:
        if conexion and conexion.is_connected():
            conexion.close()
            print( "Conexión a la base de datos cerrada.")

def procesar_fiscal_gde(filepath):
    """
    Procesa un archivo fiscal CSV (tabulado) y genera (yields) mensajes de progreso.
    """
    conexion = None # Inicializa la conexión fuera del try para el finally
    try:
        # Intenta conectar a la base de datos primero
        print( "Conectando a la base de datos...")
        conexion = connect()
        if not conexion or not conexion.is_connected():
            print( "Error: No se pudo conectar a la base de datos.")
            return # Termina la ejecución si no hay conexión
        
        print( "Abriendo archivo CSV...")
        df = pd.read_csv(filepath, sep="\t", encoding='latin1')  # usa tabulador como separador
        total_rows = len(df)
        print( colored(f"{OK_SYMBOL}Archivo abierto. {total_rows} filas encontradas.","green"))
        existe = 0
        nuevo = 0
        for idx, row in df.iterrows():
            try:
                # Validar campos requeridos
                if pd.isnull(row["Tipo DTE"]) or pd.isnull(row["Folio"]) or pd.isnull(row["RUT Cliente"]):
                    print( f"Advertencia: Saltando fila {idx+2} por datos faltantes obligatorios.")
                    continue

                fecha_latin = row["Emisión"]
                fecha = convertir_fecha_latin_a_iso(fecha_latin)
                tipo = int(row["Tipo DTE"])
                folio = int(row["Folio"]) if pd.notnull(row["Folio"]) else 0
                anticipo = ''
                conf_no = ''
                razon_social = row["Cliente"] if pd.notnull(row["Cliente"]) else 'S/Razón Social'
                rut = formatRut(str(row["RUT Cliente"])) if pd.notnull(row["RUT Cliente"]) else ''
                neto = row["Monto Neto"] if pd.notnull(row["Monto Neto"]) else 0
                exento = row["Monto Exento"] if pd.notnull(row["Monto Exento"]) else 0
                iva = row["IVA"] if pd.notnull(row["IVA"]) else 0
                total = row["Monto Total"] if pd.notnull(row["Monto Total"]) else 0
                cajero = row["Nombre POS"] if pd.notnull(row["Nombre POS"]) else ''
                index = f"{tipo}_{folio}"

                print( colored(f"\n{NEW_SYMBOL} Procesando fila {idx+1}/{total_rows}: DTE {tipo}-{folio}, Fecha: {fecha_latin}, RUT {rut}, Razon Social: {razon_social} por monto Total: {total}", "blue"))


                

                #consultando si el dte existe primero en nuestra base antes de intentar insertarlo
                print( f" {ASK_SYMBOL} -> Consultando si el DTE ya existe en la base de datos...")
                try:
                    
                    query = f"""
                                SELECT COUNT(*) FROM fiscales WHERE tipo = {tipo} AND folio = {folio}
                            """
                    cursor = conexion.cursor()
                    cursor.execute(query)
                    result = cursor.fetchone()
                    cursor.close()
                

                    if result[0] > 0:
                        print( colored(f" {WARNING_SYMBOL} -> DTE {tipo}-{folio} ya existe en la base de datos. Saltando inserción.", "yellow"))
                        existe += 1
                        continue
                        
                    else:
                        print( colored(f" {OK_SYMBOL} -> DTE {tipo}-{folio} no existe en la base de datos. Procediendo a insertar.", "green"))
                        nuevo += 1
                        # Consultar estado SII
                        print( colored(f" {ASK_SYMBOL} -> Consultando estado SII para {tipo}-{folio}...", "white"))
                        try:
                            estadoSii = consultaEstado(tipo, folio, rutEmisor)['Data']
                            print( colored(f" {OK_SYMBOL} -> Estado SII: {estadoSii}","green"))
                        except Exception as e:
                            print( colored(f" {ERROR_SYMBOL} -> Error al consultar estado SII: {e}", "red"))
                            estadoSii = None
                        
                        
                        # Consultar PDF
                        print(f" {ASK_SYMBOL} -> Consultando PDF para {tipo}-{folio}...")
                        try:
                            pdf_response = consultaPDF(tipo, folio, rutEmisor)
                            pdf_data = pdf_response.get('Data', 'Error en respuesta API')
                            print( colored(f" {OK_SYMBOL} -> PDF obtenido (PDF Base64 encoded: {str(pdf_data)[:50]}...)", "green"))
                        except Exception as e:
                            print( colored(f" {ERROR_SYMBOL} -> Error al recuperar PDF: {e}", "red"))
                            pdf_data = None

                        # Consultar XML

                        
                        print( f" {ASK_SYMBOL} -> Consultando XML para {tipo}-{folio}...")
                        try:
                            payload_xml = {
                                            "Environment": "P",
                                            "Group": "E",
                                            "Rut": rutEmisor,
                                            "DocType": tipo,
                                            "Folio": folio,
                                            "IsForDistribution": "false"
                                            }
                            xml_data = recuperar_xml(payload_xml)['RecoverDocumentResult']['Data']
                            print( colored(f" {OK_SYMBOL} -> XML obtenido (XML Base64 ecoded: {str(xml_data)[:50]}...)", "green"))
                        except Exception as e:
                            print( colored(f" {ERROR_SYMBOL} -> Error al recuperar XML: {e}", "red"))
                            xml_data = None



                        query = f"""
                            INSERT IGNORE INTO fiscales
                            (`fecha`, `tipo`, `folio`, `anticipo`, `conf_no`, `razon_social`, `rut`, `neto`, `exento`, `iva`, `total`, `cajero`, `index`, `estadoSii`, `pdf`, `xml`)
                            VALUES
                            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """
                        # Usar parámetros previene inyección SQL y maneja tipos de datos
                        params = (
                            fecha, tipo, folio, anticipo, conf_no, razon_social, rut,
                            neto, exento, iva, total, cajero, index, estadoSii, pdf_data, xml_data
                        )

                        #print(params)
                        #input("Presione una tecla para continuar...")

                        print( colored(f" {NEW_SYMBOL} -> Insertando DTE {tipo}-{folio} en la base de datos...", "blue"))
                        try:
                            cursor = conexion.cursor()
                            cursor.execute(query, params)
                            conexion.commit()
                            #rows_affected = cursor.rowcount
                            cursor.close()
                            print(colored( f"  {OK_SYMBOL} -> DTE {tipo}-{folio} insertado correctamente.", "green"))
                        except mysql.connector.Error as db_err:
                            print( colored(f"  {ERROR_SYMBOL} -> Error de Base de Datos insertando {tipo}-{folio}: {db_err}", "red"))
                            conexion.rollback()
                except mysql.connector.Error as db_err:
                    print( colored(f" {ERROR_SYMBOL} -> Error de Base de Datos consultando {tipo}-{folio}: {db_err}", "red"))
                continue


            except requests.exceptions.RequestException as api_err:
                print( colored(f" {ERROR_SYMBOL} -> Error de API consultando {tipo}-{folio}: {api_err}", "red"))
            except Exception as e:
                print( colored(f" {ERROR_SYMBOL} -> Error inesperado procesando fila {idx+2} (DTE {tipo}-{folio}): {e}", "red"))

        print( colored(f" {OK_SYMBOL} Procesamiento completado. {total_rows} filas leídas del archivo; registros existente : {existe}, registros nuevos: {nuevo}", "green"))

    except FileNotFoundError:
        print( f"Error: Archivo no encontrado.")
    except mysql.connector.Error as err:
        print( f"Error de conexión inicial a la base de datos: {err}")

    except pd.errors.ParserError:
        print( "Error: El archivo no se pudo leer correctamente. ¿Está separado por tabulaciones?")
    except Exception as e:
        print( f"Error general durante el procesamiento: {e}")


